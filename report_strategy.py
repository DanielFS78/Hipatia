# =================================================================================
# report_strategy.py
# Implementa el patrón de diseño Strategy para la generación de informes.
# =================================================================================

import logging
from abc import ABC, abstractmethod
from datetime import datetime, timedelta, time, date
from collections import defaultdict
import pandas as pd
from typing import List, Dict
# Openpyxl para gráficos y formato avanzado en Excel
# Openpyxl para gráficos y formato avanzado en Excel
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ReportLab para gráficos y maquetación en PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Spacer, Paragraph, Table, TableStyle, PageBreak
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.axes import XCategoryAxis, YValueAxis

# Módulos de la aplicación
from time_calculator import CalculadorDeTiempos
from calculation_audit import CalculationDecision, DecisionStatus  # Importamos el modelo de auditoría

class IReporteEstrategia(ABC):
    @abstractmethod
    def generar_reporte(self, datos_informe, output_path) -> bool:
        pass

class ReportePilaFabricacionExcelMejorado(IReporteEstrategia):
    """
    Generador mejorado de reportes Excel con lectura correcta del audit_log
    y presentación clara de grupos secuenciales.
    """

    def __init__(self, schedule_config=None):
        super().__init__()
        self.logger = logging.getLogger(__name__)
        self.schedule_config = schedule_config
        self.time_calculator = CalculadorDeTiempos(self.schedule_config) if schedule_config else None
        self.workbook = None

    def generar_reporte(self, datos_informe: dict) -> bool:
        """
        Orquesta la creación de todas las hojas del informe en memoria.
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Eliminar la hoja por defecto
            # Guardar datos_informe en el workbook para acceso desde otras funciones
            wb._datos_informe = datos_informe

            all_results = datos_informe.get("data", [])
            audit_log = datos_informe.get("audit_log", [])
            production_flow = datos_informe.get("production_flow", [])

            if not all_results:
                self.logger.warning("No hay datos de simulación para generar el reporte Excel.")
                return False

            self.logger.info(f"Iniciando generación de informe con {len(all_results)} registros cronológicos.")

            # COMIENZA A COPIAR DESDE AQUÍ (Insertar después de la línea anterior)

            # --- INICIO DIAGNÓSTICO CUELLOS DE BOTELLA ---
            self.logger.info(f"DEBUG: Contenido de audit_log recibido ({len(audit_log)} eventos):")
            if not audit_log:
                self.logger.warning("DEBUG: ¡El audit_log está vacío!")
            else:
                # Mostrar los tipos de decisión y estados de los primeros 5 eventos para inspección
                for i, event in enumerate(audit_log[:5]):
                    if isinstance(event, CalculationDecision):
                        self.logger.info(
                            f"  Evento {i}: Tipo='{event.decision_type}', Estado='{event.status.value if hasattr(event.status, 'value') else event.status}', Razón='{event.user_friendly_reason[:50]}...'")
                    else:
                        self.logger.info(f"  Evento {i}: (Tipo no reconocido: {type(event)})")
            # --- FIN DIAGNÓSTICO ---

            # TERMINA DE COPIAR AQUÍ

            # 1. Analizar los datos completos una sola vez
            analysis = self._analyze_simulation_data(all_results, audit_log)

            # 2. Crear cada hoja del reporte llamando a los métodos de maquetación
            self._crear_hoja_resumen_ejecutivo(wb, analysis, datos_informe)
            self._crear_hoja_analisis_trabajadores(wb, all_results, audit_log)
            self._crear_hoja_graficas(wb, all_results, analysis)
            self._crear_hoja_cronograma(wb, all_results) # Se pasan todos los resultados ordenados
            self._crear_hoja_cuellos_botella(wb, audit_log)
            self._crear_hoja_trabajo_paralelo(wb, all_results)
            self._crear_hoja_audit_detallado(wb, audit_log)

            self.workbook = wb
            self.logger.info("Informe Excel generado en memoria con todas las hojas y maquetación.")
            return True

        except Exception as e:
            self.logger.error(f"Error crítico durante la generación del reporte Excel: {e}", exc_info=True)
            return False

    def _agrupar_eventos_relacionados(self, eventos: List[CalculationDecision], umbral_segundos: int = 5) -> List[Dict]:
        """
        Agrupa eventos de auditoría que ocurren en la misma tarea y en un
        intervalo de tiempo cercano.
        (Basado en la Propuesta 6.3 del plan)

        Args:
            eventos: Lista de objetos CalculationDecision (ya filtrados si es necesario).
            umbral_segundos: Máxima diferencia en segundos para considerar eventos como relacionados.

        Returns:
            Lista de diccionarios, donde cada diccionario representa un grupo:
            [{'tarea': 'NombreTarea', 'eventos': [evento1, evento2]}, ...]
        """
        if not eventos:
            return []

        # Asegurarse de que los eventos estén ordenados por timestamp
        eventos_ordenados = sorted(eventos, key=lambda x: x.timestamp if hasattr(x,
                                                                                 'timestamp') and x.timestamp else datetime.min)

        grupos = []
        grupo_actual = None

        umbral_delta = timedelta(seconds=umbral_segundos)

        for evento in eventos_ordenados:
            # Necesitamos un timestamp válido para agrupar
            if not hasattr(evento, 'timestamp') or not evento.timestamp:
                # Si un evento no tiene timestamp, lo añadimos a un grupo genérico o lo saltamos
                # Por simplicidad, lo añadimos al último grupo si existe, o creamos uno nuevo
                if grupo_actual:
                    grupo_actual['eventos'].append(evento)
                else:
                    grupos.append({'tarea': evento.task_name or 'Desconocida', 'eventos': [evento]})
                continue

            # Si no hay grupo actual, o el evento es de otra tarea, o ha pasado mucho tiempo
            if (grupo_actual is None or
                    evento.task_name != grupo_actual['tarea'] or
                    (evento.timestamp - grupo_actual['timestamp_ultimo']) > umbral_delta):

                # Crear un nuevo grupo
                grupo_actual = {
                    'tarea': evento.task_name or 'Desconocida',
                    'timestamp_ultimo': evento.timestamp,
                    'eventos': [evento]
                }
                grupos.append(grupo_actual)
            else:
                # Añadir al grupo actual
                grupo_actual['eventos'].append(evento)
                # Actualizar el timestamp del último evento del grupo
                grupo_actual['timestamp_ultimo'] = evento.timestamp

        self.logger.info(f"Eventos de auditoría agrupados en {len(grupos)} bloques.")
        return grupos

    def _crear_hoja_audit_detallado(self, wb, audit_log, hay_limite=False, total_original=None):
        """
        Crea una hoja con el log de auditoría detallado, filtrado y agrupado.
        MEJORADO: Filtra eventos neutros y agrupa visualmente .
        """
        ws = wb.create_sheet("Audit Log")

        # Título
        ws['A1'] = "LOG DE AUDITORÍA (Eventos Relevantes Agrupados)"  # Título modificado
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:F1')

        # Advertencia si hay límite
        if hay_limite and total_original:
            ws[
                'A2'] = f"⚠️ NOTA: Mostrando {len(audit_log)} de {total_original} eventos totales (límite por optimización)"
            ws['A2'].font = Font(size=10, italic=True, color="FF0000")
            ws.merge_cells('A2:F2')
            header_row = 4
        else:
            header_row = 3

        # Encabezados (sin cambios)
        headers = ["Timestamp", "Tarea", "Tipo de Decisión", "Descripción", "Estado", "Producto Asociado"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row = header_row + 1

        # --- INICIO: Filtrado y Agrupación ---
        # 1. Filtrar eventos importantes
        eventos_importantes = []
        tipos_relacionados_grupo = ['GRUPO_SECUENCIAL', 'INSTANCIA_PARALELA']  # Añadir tipos relacionados si existen

        for evento in audit_log:
            # Asegurarse de que es un objeto CalculationDecision válido
            if not isinstance(evento, CalculationDecision):
                continue

            # Comprobar si el evento es importante
            es_importante = (
                    evento.status != DecisionStatus.NEUTRAL or
                    any(tipo in evento.decision_type for tipo in tipos_relacionados_grupo)
            )

            if es_importante:
                eventos_importantes.append(evento)

        # 2. Agrupar los eventos filtrados
        grupos_eventos = self._agrupar_eventos_relacionados(eventos_importantes)
        # --- FIN: Filtrado y Agrupación ---

        # 3. Iterar sobre los grupos para escribir en la hoja
        if not grupos_eventos:
            ws.cell(row=current_row, column=1, value="No hay eventos relevantes que mostrar.")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
        else:
            for grupo in grupos_eventos:
                # Escribir filas para cada evento dentro del grupo
                for event in grupo['eventos']:
                    # Extraer datos del evento (igual que antes)
                    timestamp_str = event.timestamp.strftime('%d/%m/%Y %H:%M:%S') if isinstance(event.timestamp,
                                                                                                datetime) else str(
                        event.timestamp)
                    product_info = "N/A"
                    if hasattr(event, 'product_code') and event.product_code:
                        product_info = event.product_code
                        if hasattr(event, 'product_desc') and event.product_desc:
                            product_info += f" / {event.product_desc}"

                    row_data = [
                        timestamp_str,
                        event.task_name,
                        event.decision_type,
                        event.user_friendly_reason,
                        event.status.value if hasattr(event.status, 'value') else str(event.status),
                        product_info
                    ]

                    # Escribir celdas y aplicar formato de estado
                    for col_num, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.value = value
                        cell.alignment = Alignment(vertical="center", wrap_text=True)  # Ajuste vertical y wrap
                        if col_num in [1, 5]:  # Timestamp y Estado centrados
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        else:  # Otros alineados a la izquierda
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                        # Formato de color por estado (solo en columna Estado)
                        if col_num == 5:
                            status_color = None
                            if event.status == DecisionStatus.POSITIVE:
                                status_color = "C6EFCE"
                            elif event.status == DecisionStatus.WARNING:
                                status_color = "FFEB9C"
                            elif event.status == DecisionStatus.CRITICAL:
                                status_color = "FFC7CE"  # Añadir CRITICAL
                            # NEUTRAL se queda sin color de fondo especial

                            if status_color:
                                cell.fill = PatternFill(start_color=status_color, end_color=status_color,
                                                        fill_type="solid")

                    current_row += 1

                # --- Añadir Separador Visual entre Grupos ---
                # Aplicar un borde inferior grueso a la última fila del grupo
                last_row_of_group = current_row - 1
                if last_row_of_group >= header_row + 1:  # Asegurar que no es la fila de cabecera
                    for col_idx in range(1, len(headers) + 1):
                        cell_sep = ws.cell(row=last_row_of_group, column=col_idx)
                        # Conservar bordes existentes y añadir el inferior
                        current_border = cell_sep.border
                        cell_sep.border = Border(left=current_border.left, right=current_border.right,
                                                 top=current_border.top, bottom=Side(style='medium', color='A0A0A0'))
                # Opcional: añadir una fila en blanco como separador adicional
                # current_row += 1
                # --- Fin Separador ---

        # Auto-ajustar columnas (sin cambios)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 60  # Aumentar ancho descripción
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 45

        # Filtros (aplicar al rango real de datos escritos)
        if current_row > header_row + 1:
            ws.auto_filter.ref = f"A{header_row}:F{current_row - 1}"
        else:  # Si no hay datos, quitar referencia al filtro
            ws.auto_filter.ref = None

        self.logger.info(
            f"Hoja de audit log creada con {len(eventos_importantes)} eventos relevantes agrupados.")  # Mensaje actualizado

    def _crear_hoja_graficas(self, wb, all_results, analysis):
        """
        Crea una hoja dedicada exclusivamente a las gráficas, organizadas de forma limpia.
        Centraliza todas las visualizaciones del informe.
        """
        ws = wb.create_sheet("📊 Gráficas")

        # Título principal
        ws['A1'] = "VISUALIZACIONES Y GRÁFICAS DEL ANÁLISIS"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:P1')

        current_row = 3

        # ========================================================================
        # GRÁFICA 1: DISTRIBUCIÓN POR DEPARTAMENTO (Circular)
        # ========================================================================
        if analysis.get('departments'):
            ws[f'A{current_row}'] = "1. DISTRIBUCIÓN DE TIEMPO POR DEPARTAMENTO"
            ws[f'A{current_row}'].font = Font(size=12, bold=True, color="4472C4")
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1

            # Datos para el gráfico
            sorted_depts = sorted(analysis['departments'].items(), key=lambda x: x[1], reverse=True)
            data_start_row = current_row

            ws[f'A{data_start_row}'] = "Departamento"
            ws[f'B{data_start_row}'] = "Tiempo (min)"
            ws[f'A{data_start_row}'].font = Font(bold=True)
            ws[f'B{data_start_row}'].font = Font(bold=True)
            data_start_row += 1

            for dept, minutes in sorted_depts:
                ws[f'A{data_start_row}'] = dept
                ws[f'B{data_start_row}'] = round(minutes, 1)
                data_start_row += 1

            data_end_row = data_start_row - 1

            # Crear gráfico circular
            chart_pie = PieChart()
            chart_pie.title = "Distribución de Tiempo por Departamento"
            chart_pie.style = 10
            chart_pie.height = 12
            chart_pie.width = 16

            data_ref = Reference(ws, min_col=2, min_row=current_row, max_row=data_end_row)
            labels_ref = Reference(ws, min_col=1, min_row=current_row + 1, max_row=data_end_row)

            chart_pie.add_data(data_ref, titles_from_data=True)
            chart_pie.set_categories(labels_ref)
            chart_pie.dataLabels = DataLabelList()
            chart_pie.dataLabels.showVal = True
            chart_pie.dataLabels.showPercent = True

            # Posicionar el gráfico
            ws.add_chart(chart_pie, f'D{current_row}')

            current_row = data_end_row + 18  # Espacio generoso después del gráfico

        # ========================================================================
        # GRÁFICA 2: TIEMPO DE PRODUCCIÓN POR PRODUCTO (Barras)
        # ========================================================================
        ws[f'A{current_row}'] = "2. TIEMPO DE PRODUCCIÓN POR PRODUCTO (TOP 10)"
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="4472C4")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        # Calcular tiempo total por producto
        product_times = defaultdict(float)
        if all_results:
            for task in all_results:
                producto_codigo = task.get('Codigo Producto', 'N/A')
                producto_key = producto_codigo if producto_codigo != 'N/A' else 'Sin Código'
                duracion = float(task.get('Duracion (min)', 0))
                product_times[producto_key] += duracion

        sorted_products = sorted(product_times.items(), key=lambda x: x[1], reverse=True)[:10]

        if sorted_products:
            data_start_row = current_row

            ws[f'A{data_start_row}'] = "Producto"
            ws[f'B{data_start_row}'] = "Tiempo (min)"
            ws[f'A{data_start_row}'].font = Font(bold=True)
            ws[f'B{data_start_row}'].font = Font(bold=True)
            data_start_row += 1

            for producto, tiempo_min in sorted_products:
                ws[f'A{data_start_row}'] = producto
                ws[f'B{data_start_row}'] = round(tiempo_min, 1)
                data_start_row += 1

            data_end_row = data_start_row - 1

            # Crear gráfico de barras
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.style = 10
            bar_chart.title = "Tiempo de Producción por Producto"
            bar_chart.y_axis.title = 'Minutos'
            bar_chart.x_axis.title = 'Productos'
            bar_chart.height = 12
            bar_chart.width = 20

            data_bar = Reference(ws, min_col=2, min_row=current_row, max_row=data_end_row)
            cats_bar = Reference(ws, min_col=1, min_row=current_row + 1, max_row=data_end_row)

            bar_chart.add_data(data_bar, titles_from_data=True)
            bar_chart.set_categories(cats_bar)
            bar_chart.dataLabels = DataLabelList()
            bar_chart.dataLabels.showVal = True

            # Posicionar el gráfico
            ws.add_chart(bar_chart, f'D{current_row}')

            current_row = data_end_row + 20

        # ========================================================================
        # GRÁFICA 3: TIEMPO TOTAL POR TRABAJADOR (Barras)
        # ========================================================================
        ws[f'A{current_row}'] = "3. CARGA DE TRABAJO POR TRABAJADOR"
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="4472C4")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        # Calcular estadísticas por trabajador
        worker_stats = defaultdict(lambda: {'tasks': 0, 'total_time': 0.0})
        for task in all_results:
            trabajadores = task.get('Trabajador Asignado', [])
            duracion = float(task.get('Duracion (min)', 0))

            if isinstance(trabajadores, list) and trabajadores:
                duracion_por_trabajador = duracion / len(trabajadores)
                for trabajador in trabajadores:
                    worker_stats[trabajador]['tasks'] += 1
                    worker_stats[trabajador]['total_time'] += duracion_por_trabajador
            elif isinstance(trabajadores, str):
                if ',' in trabajadores:
                    trabajadores_separados = [t.strip() for t in trabajadores.split(',')]
                    duracion_por_trabajador = duracion / len(trabajadores_separados)
                    for trabajador in trabajadores_separados:
                        worker_stats[trabajador]['tasks'] += 1
                        worker_stats[trabajador]['total_time'] += duracion_por_trabajador
                else:
                    worker_stats[trabajadores]['tasks'] += 1
                    worker_stats[trabajadores]['total_time'] += duracion

        sorted_workers = sorted(worker_stats.items(), key=lambda x: x[1]['total_time'], reverse=True)

        if sorted_workers:
            data_start_row = current_row

            ws[f'A{data_start_row}'] = "Trabajador"
            ws[f'B{data_start_row}'] = "Tiempo Total (min)"
            ws[f'A{data_start_row}'].font = Font(bold=True)
            ws[f'B{data_start_row}'].font = Font(bold=True)
            data_start_row += 1

            for trabajador, stats in sorted_workers:
                ws[f'A{data_start_row}'] = trabajador
                ws[f'B{data_start_row}'] = round(stats['total_time'], 1)
                data_start_row += 1

            data_end_row = data_start_row - 1

            # Crear gráfico de barras
            worker_chart = BarChart()
            worker_chart.type = "col"
            worker_chart.style = 10
            worker_chart.title = "Tiempo Total por Trabajador"
            worker_chart.y_axis.title = 'Minutos'
            worker_chart.x_axis.title = 'Trabajadores'
            worker_chart.height = 12
            worker_chart.width = 20

            data_worker = Reference(ws, min_col=2, min_row=current_row, max_row=data_end_row)
            cats_worker = Reference(ws, min_col=1, min_row=current_row + 1, max_row=data_end_row)

            worker_chart.add_data(data_worker, titles_from_data=True)
            worker_chart.set_categories(cats_worker)
            worker_chart.dataLabels = DataLabelList()
            worker_chart.dataLabels.showVal = True

            # Posicionar el gráfico
            ws.add_chart(worker_chart, f'D{current_row}')

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

        self.logger.info("✅ Hoja de gráficas creada con 3 visualizaciones")

    def _crear_hoja_analisis_trabajadores(self, wb, all_results, audit_log):
        """
        Crea una hoja con análisis detallado por trabajador.
        CORREGIDO: Añade la columna de jornadas laborales.
        """
        ws = wb.create_sheet("Análisis Trabajadores")

        # Título
        ws['A1'] = "ANÁLISIS POR TRABAJADOR"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        # --- MODIFICACIÓN: Ampliar el título a la nueva columna G ---
        ws.merge_cells('A1:G1')

        # Encabezados
        row = 3
        # --- MODIFICACIÓN: Añadir nuevo encabezado ---
        headers = ["Trabajador", "Tareas Asignadas", "Tiempo Total (min)", "Tiempo Total (horas)",
                   "Jornadas Laborales", "Tiempo Promedio/Tarea", "Carga (%)"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Agrupar datos por trabajador (sin cambios en esta parte)
        worker_stats = defaultdict(lambda: {'tasks': 0, 'total_time': 0.0})
        for task in all_results:
            trabajadores = task.get('Trabajador Asignado', [])
            # ✅ AÑADE ESTA LÍNEA DE DEBUG:
            self.logger.info(
                f"DEBUG: Trabajador={trabajadores}, Tipo={type(trabajadores)}, Tarea={task.get('Tarea', '?')[:30]}")
            duracion = float(task.get('Duracion (min)', 0))
            duracion = float(task.get('Duracion (min)', 0))
            if isinstance(trabajadores, list) and trabajadores:
                duracion_por_trabajador = duracion / len(trabajadores)
                for trabajador in trabajadores:
                    worker_stats[trabajador]['tasks'] += 1
                    worker_stats[trabajador]['total_time'] += duracion_por_trabajador
            elif isinstance(trabajadores, str):
                # Si es un string con múltiples trabajadores separados por coma
                if ',' in trabajadores:
                    trabajadores_separados = [t.strip() for t in trabajadores.split(',')]
                    duracion_por_trabajador = duracion / len(trabajadores_separados)
                    for trabajador in trabajadores_separados:
                        worker_stats[trabajador]['tasks'] += 1
                        worker_stats[trabajador]['total_time'] += duracion_por_trabajador
                else:
                    # Es un solo trabajador
                    worker_stats[trabajadores]['tasks'] += 1
                    worker_stats[trabajadores]['total_time'] += duracion

        max_time = max((stats['total_time'] for stats in worker_stats.values()), default=1)
        sorted_workers = sorted(worker_stats.items(), key=lambda x: x[1]['total_time'], reverse=True)

        for trabajador, stats in sorted_workers:
            tiempo_min = stats['total_time']
            tiempo_horas = tiempo_min / 60
            num_tareas = stats['tasks']
            tiempo_promedio = tiempo_min / num_tareas if num_tareas > 0 else 0
            carga_porcentaje = (tiempo_min / max_time * 100) if max_time > 0 else 0

            # --- INICIO DE LA MODIFICACIÓN ---
            # Calcular las jornadas laborales asumiendo 8 horas/día (480 min)
            jornadas = tiempo_min / 480
            # --- FIN DE LA MODIFICACIÓN ---

            ws[f'A{row}'] = trabajador
            ws[f'B{row}'] = num_tareas
            ws[f'C{row}'] = round(tiempo_min, 1)
            ws[f'D{row}'] = round(tiempo_horas, 2)
            # --- MODIFICACIÓN: Escribir el nuevo dato en la columna E ---
            ws[f'E{row}'] = round(jornadas, 2)
            ws[f'F{row}'] = round(tiempo_promedio, 1)
            ws[f'G{row}'] = f"{carga_porcentaje:.1f}%"

            # Formato condicional para carga (ahora en columna G)
            if carga_porcentaje > 80:
                ws[f'G{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif carga_porcentaje > 60:
                ws[f'G{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                ws[f'G{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            row += 1

        # Auto-ajustar columnas
        # --- MODIFICACIÓN: Añadir la nueva columna G ---
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18

        # Gráfico de barras
        if len(sorted_workers) > 0:
            chart = BarChart()
            chart.title = "Tiempo Total por Trabajador (minutos)"
            chart.style = 10
            chart.x_axis.title = "Trabajador"
            chart.y_axis.title = "Minutos"

            data_ref = Reference(ws, min_col=3, min_row=3, max_row=row - 1)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

        # === NUEVA SECCIÓN: DESGLOSE DETALLADO POR TRABAJADOR ===
        row += 2  # Espacio después del resumen

        ws[f'A{row}'] = "=== DESGLOSE DETALLADO POR TRABAJADOR ==="
        ws[f'A{row}'].font = Font(size=13, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(f'A{row}:G{row}')

        row += 2

        # Agrupar datos por trabajador y tarea
        trabajador_tarea_unidades = defaultdict(lambda: defaultdict(list))

        for task in all_results:
            trabajadores = task.get('Trabajador Asignado', [])
            if not isinstance(trabajadores, list):
                if isinstance(trabajadores, str):
                    if ',' in trabajadores:
                        trabajadores = [t.strip() for t in trabajadores.split(',')]
                    else:
                        trabajadores = [trabajadores]
                else:
                    trabajadores = []

            for trabajador in trabajadores:
                if not trabajador or not str(trabajador).strip():
                    continue

                # COMIENZA A COPIAR DESDE AQUÍ
                tarea_nombre = task.get('Tarea', 'N/A')

                # --- INICIO CORRECCIÓN: Obtener producto correctamente ---
                product_code = task.get('Codigo Producto', '')
                product_desc = task.get('Descripcion Producto', '')
                producto_str = f"{product_code} - {product_desc}" if product_desc else product_code or 'N/A'
                # --- FIN CORRECCIÓN ---

                trabajador_tarea_unidades[trabajador][tarea_nombre].append({
                    'unidad': task.get('Numero Unidad', '?'),
                    'fin': task.get('Fin'),
                    'duracion': task.get('Duracion (min)', 0),
                    'producto': producto_str  # <-- Usar la cadena construida
                })
                # TERMINA DE COPIAR AQUÍ

        # Ordenar trabajadores alfabéticamente
        for trabajador in sorted(trabajador_tarea_unidades.keys()):
            # Header del trabajador
            ws[f'A{row}'] = f"👤 {trabajador}"
            ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.merge_cells(f'A{row}:G{row}')
            row += 1

            # Por cada tarea de este trabajador
            for tarea_nombre, unidades_list in sorted(trabajador_tarea_unidades[trabajador].items()):
                # Sub-header de la tarea
                ws[f'A{row}'] = f"📋 Tarea: {tarea_nombre} ({len(unidades_list)} unidades)"
                ws[f'A{row}'].font = Font(size=11, bold=True, italic=True)
                ws[f'A{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                ws.merge_cells(f'A{row}:G{row}')
                row += 1

                # Headers de la tabla de unidades
                headers_unidad = ["Unidad #", "Finalizada", "Duración (min)", "Producto"]
                for col_num, header in enumerate(headers_unidad, start=2):  # Empezar en B
                    cell = ws.cell(row=row, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True, size=9)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                row += 1

                # Ordenar unidades por timestamp
                unidades_ordenadas = sorted(unidades_list, key=lambda x: x['fin'] if x['fin'] else datetime.min)

                # Datos de cada unidad
                for unidad_data in unidades_ordenadas:
                    ws.cell(row, 2, unidad_data['unidad'])  # Columna B
                    ws.cell(row, 3,
                            unidad_data['fin'].strftime('%d/%m/%Y %H:%M') if unidad_data['fin'] else 'N/A')  # C
                    ws.cell(row, 4, f"{unidad_data['duracion']:.1f}")  # D
                    ws.cell(row, 5, unidad_data['producto'])  # E

                    # Alineación
                    ws.cell(row, 2).alignment = Alignment(horizontal="center")
                    ws.cell(row, 3).alignment = Alignment(horizontal="center")
                    ws.cell(row, 4).alignment = Alignment(horizontal="center")
                    ws.cell(row, 5).alignment = Alignment(horizontal="left")

                    row += 1

                row += 1  # Espacio entre tareas

            row += 1  # Espacio entre trabajadores

        self.logger.info(f"Hoja de análisis de trabajadores creada con {len(worker_stats)} trabajadores")

    def _analyze_simulation_data(self, results, audit_log):
        """
        Analiza los datos de simulación para extraer métricas clave.
        """
        analysis = {
            'total_tasks': len(results),
            'total_duration_min': sum(r['Duracion (min)'] for r in results),
            'start_time': min(r['Inicio'] for r in results) if results else None,
            'end_time': max(r['Fin'] for r in results) if results else None,
            'workers_involved': set(),
            'machines_used': set(),
            'departments': defaultdict(float),
            'idle_times': [],
            'bottlenecks': [],
            'groups_performance': []
        }

        # Analizar trabajadores y máquinas
        for result in results:
            workers = result.get('Trabajador Asignado', [])

            # Logging para debugging
            self.logger.debug(f"Procesando trabajadores: {workers} (tipo: {type(workers)})")

            if isinstance(workers, list):
                # Filtrar valores vacíos, None o strings vacíos
                valid_workers = [w for w in workers if w and str(w).strip()]
                analysis['workers_involved'].update(valid_workers)
            elif isinstance(workers, str):
                # Manejar strings con múltiples trabajadores separados por coma
                if ',' in workers:
                    worker_list = [w.strip() for w in workers.split(',') if w.strip()]
                    analysis['workers_involved'].update(worker_list)
                elif workers and workers.strip():
                    # Trabajador único (no vacío)
                    analysis['workers_involved'].add(workers.strip())
            # Si workers es None o vacío, no se añade nada

            machine = result.get('nombre_maquina')
            if machine and machine != 'N/A':
                analysis['machines_used'].add(machine)

            dept = result.get('Departamento', 'General')
            analysis['departments'][dept] += result['Duracion (min)']

        # Log final para verificar el conteo correcto de trabajadores
        self.logger.info(f"✅ Total trabajadores identificados: {len(analysis['workers_involved'])}")
        self.logger.info(f"📋 Lista de trabajadores: {sorted(analysis['workers_involved'])}")

        # Analizar tiempos muertos del audit_log

        # Analizar tiempos muertos del audit_log
        for i, decision in enumerate(audit_log):
            if hasattr(decision, 'decision_type'):
                if decision.decision_type == 'TIEMPO_DE_ESPERA':
                    analysis['idle_times'].append({
                        'task': decision.task_name,
                        'duration': decision.details.get('wait_time', 0),
                        'reason': decision.reason
                    })
                elif decision.decision_type == 'CUELLO_DE_BOTELLA':
                    analysis['bottlenecks'].append({
                        'resource': decision.details.get('resource'),
                        'impact': decision.details.get('impact_minutes', 0),
                        'affected_tasks': decision.details.get('affected_tasks', [])
                    })

                # --- INICIO: PASO 6.1.1 - Calcular Métricas de Paralelismo ---
                self.logger.debug("Analizando métricas de trabajo paralelo...")
                instancias_encontradas = set()
                tareas_con_paralelo = defaultdict(set)

                for task in results:  # 'results' es all_results
                    inst_id = task.get('Instancia ID')
                    # Nos aseguramos de que sea un ID válido y no el principal/N/A
                    if inst_id and inst_id != 'N/A' and inst_id != 'Principal':
                        instancias_encontradas.add(inst_id)

                        tarea_nombre = task.get('Tarea', 'Desconocida')
                        tareas_con_paralelo[tarea_nombre].add(inst_id)

                # Contar solo las instancias que son parte de un paralelismo (>1 por tarea)
                total_instancias_paralelas = 0
                max_instancias_simultaneas_en_tarea = 0

                for tarea, instancias_set in tareas_con_paralelo.items():
                    num_instancias = len(instancias_set)
                    # Solo contamos si la tarea realmente tuvo paralelismo
                    if num_instancias > 1:
                        # Sumamos el total de instancias involucradas en paralelismo
                        total_instancias_paralelas += num_instancias
                        # Buscamos la tarea con más instancias simultáneas
                        if num_instancias > max_instancias_simultaneas_en_tarea:
                            max_instancias_simultaneas_en_tarea = num_instancias

                # Si no hubo tareas con >1 instancia, pero sí se usaron instancias,
                # puede que el log esté separado (ej. 1 instancia 'Principal', 1 'paralela')
                # Por seguridad, si el cálculo anterior dio 0 pero hay instancias,
                # mostramos el recuento total.
                if total_instancias_paralelas == 0 and len(instancias_encontradas) > 0:
                    total_instancias_paralelas = len(instancias_encontradas)

                if max_instancias_simultaneas_en_tarea == 0 and len(instancias_encontradas) > 0:
                    max_instancias_simultaneas_en_tarea = max(
                        len(s) for s in tareas_con_paralelo.values()) if tareas_con_paralelo else 0

                analysis['total_instancias_paralelas'] = total_instancias_paralelas
                analysis['max_instancias_simultaneas'] = max_instancias_simultaneas_en_tarea
                self.logger.info(
                    f"Análisis de paralelismo: Total Instancias={total_instancias_paralelas}, Max Simultáneas={max_instancias_simultaneas_en_tarea}")
                # --- FIN: PASO 6.1.1 ---

        return analysis


    def _crear_hoja_resumen_ejecutivo(self, wb, analysis, datos_informe):
        """
        Crea una hoja de resumen ejecutivo con métricas clave y visualizaciones.
        MODIFICADO: Corregido cálculo de unidades totales y mantenido cálculo de jornadas.
        """

        ws = wb.create_sheet("Resumen Ejecutivo")

        # Título
        ws['A1'] = "RESUMEN EJECUTIVO - ANÁLISIS DE PRODUCCIÓN"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:H1')

        # Información general
        row = 3
        ws[f'A{row}'] = "INFORMACIÓN GENERAL"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:D{row}')

        row += 1

        # --- INICIO CORRECCIÓN: Cálculo correcto de unidades totales ---
        # Primero procesamos los productos para obtener las unidades
        all_results = datos_informe.get("data", [])
        productos_info = defaultdict(lambda: {
            'unidades': 0,
            'inicio': None,
            'fin': None
        })
        if all_results:  # Asegurarse de que hay datos
            for task in all_results:
                producto_codigo = task.get('Codigo Producto', 'N/A')
                producto_desc = task.get('Descripcion Producto', '')
                producto_key = f"{producto_codigo} - {producto_desc}" if producto_desc else producto_codigo

                # CORRECCIÓN: Usar 'Numero Unidad' si existe, como indicador del progreso
                # Asumimos que el número máximo de unidad indica el total para ese producto
                unidad_actual = task.get('Numero Unidad', 1)
                try:
                    unidad_actual = int(unidad_actual)
                except (ValueError, TypeError):
                    unidad_actual = 1

                if unidad_actual > productos_info[producto_key]['unidades']:
                    productos_info[producto_key][
                        'unidades'] = unidad_actual  # Guardamos el número MÁXIMO de unidad visto

                # Actualizar fechas de inicio y fin (sin cambios)
                inicio_tarea = task.get('Inicio')
                fin_tarea = task.get('Fin')

                if inicio_tarea:
                    if productos_info[producto_key]['inicio'] is None or inicio_tarea < \
                            productos_info[producto_key]['inicio']:
                        productos_info[producto_key]['inicio'] = inicio_tarea

                if fin_tarea:
                    if productos_info[producto_key]['fin'] is None or fin_tarea > productos_info[producto_key][
                        'fin']:
                        productos_info[producto_key]['fin'] = fin_tarea

            # Ahora sumamos las unidades máximas de cada producto
            unidades_totales_calculadas = sum(info['unidades'] for info in productos_info.values())
            if unidades_totales_calculadas == 0 and all_results:  # Si la suma es 0 pero hay tareas, usar el número de tareas como fallback
                unidades_totales_calculadas = len(all_results)
            elif not all_results:
                unidades_totales_calculadas = 0  # Si no hay resultados, las unidades son 0
        else:
            unidades_totales_calculadas = 0  # Si all_results está vacío

        total_tareas_individuales = len(all_results)  # Renombrado para claridad

        # Calcular el tiempo total por trabajador para encontrar el máximo (Lógica sin cambios)
        jornadas_laborales = 0
        num_trabajadores = 0  # Inicializar
        if all_results:
            worker_stats = defaultdict(float)
            unique_workers = set()  # Para contar trabajadores únicos
            for task in all_results:
                trabajadores = task.get('Trabajador Asignado', [])
                duracion = float(task.get('Duracion (min)', 0))

                # Manejo robusto de la asignación de trabajadores
                if isinstance(trabajadores, list) and trabajadores:
                    # Filtrar nombres vacíos o None
                    valid_workers = [w for w in trabajadores if w and str(w).strip()]
                    if valid_workers:
                        unique_workers.update(valid_workers)
                        duracion_por_trabajador = duracion / len(valid_workers)
                        for trabajador in valid_workers:
                            worker_stats[trabajador] += duracion_por_trabajador
                elif isinstance(trabajadores, str) and trabajadores.strip() and trabajadores != 'N/A':
                    # Si es un string con nombres separados por coma
                    if ',' in trabajadores:
                        trabajadores_list = [t.strip() for t in trabajadores.split(',') if t.strip()]
                        if trabajadores_list:
                            unique_workers.update(trabajadores_list)
                            duracion_por_trabajador = duracion / len(trabajadores_list)
                            for trabajador in trabajadores_list:
                                worker_stats[trabajador] += duracion_por_trabajador
                    else:  # Es un solo trabajador
                        unique_workers.add(trabajadores)
                        worker_stats[trabajadores] += duracion

            num_trabajadores = len(unique_workers)  # Usar el conteo de trabajadores únicos

            if worker_stats:
                max_worker_time = max(worker_stats.values())
                jornadas_laborales = max_worker_time / 480  # 480 min = 8 horas

        # Lista de información actualizada
        info_items = [
            ("Fabricación:", datos_informe.get("fab_info", "N/A")),
            ("Unidades Totales:", unidades_totales_calculadas),  # <-- VALOR CORREGIDO
            ("Trabajadores implicados:", f"{num_trabajadores} trabajadores"),  # <-- Etiqueta cambiada
            ("Total de tareas individuales:", total_tareas_individuales),  # <-- Etiqueta cambiada
            ("Jornadas laborales (trabajador + ocupado):", f"{jornadas_laborales:.1f} días"),
            # <-- Etiqueta aclarada
            ("Fecha inicio:",
                analysis['start_time'].strftime('%d/%m/%Y %H:%M') if analysis['start_time'] else "N/A"),
            ("Fecha fin:", analysis['end_time'].strftime('%d/%m/%Y %H:%M') if analysis['end_time'] else "N/A"),
            ("Duración total (tiempo productivo):",
             f"{analysis['total_duration_min']:.1f} minutos ({analysis['total_duration_min'] / 60:.1f} horas)")
        ]

        # --- INICIO: PASO 6.1.2 - Añadir Métricas de Paralelismo al Resumen ---
        total_inst = analysis.get('total_instancias_paralelas', 0)
        max_inst = analysis.get('max_instancias_simultaneas', 0)

        # Solo mostrar si se detectó paralelismo
        if total_inst > 0 or max_inst > 0:
            info_items.extend([
                ("Total Instancias Paralelas:", f"{total_inst} instancias"),
                ("Máx. Instancias Simultáneas (en 1 Tarea):", f"{max_inst} instancias")
            ])

        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # NUEVA SECCIÓN: Detalle de productos fabricados
        row += 2
        ws[f'A{row}'] = "PRODUCTOS FABRICADOS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # --- USAR productos_info YA CALCULADO ARRIBA ---
        productos_ordenados = sorted(productos_info.items())

        # Crear tabla de productos (Sin cambios en esta parte)
        if productos_ordenados:
            headers_productos = ["Producto", "Unidades", "Inicio Producción", "Fin Producción"]
            for col_num, header in enumerate(headers_productos, start=1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for producto, info in productos_ordenados:
                ws[f'A{row}'] = producto
                ws[f'B{row}'] = info['unidades']  # Ahora muestra el número MÁXIMO de unidad
                ws[f'C{row}'] = info['inicio'].strftime('%d/%m/%Y %H:%M') if info['inicio'] else "N/A"
                ws[f'D{row}'] = info['fin'].strftime('%d/%m/%Y %H:%M') if info['fin'] else "N/A"
                ws[f'B{row}'].alignment = Alignment(horizontal="center")
                ws[f'C{row}'].alignment = Alignment(horizontal="center")
                ws[f'D{row}'].alignment = Alignment(horizontal="center")
                row += 1
        else:
            ws[f'A{row}'] = "No hay información de productos disponible"
            ws[f'A{row}'].font = Font(italic=True)
            row += 1

        # Métricas de eficiencia (Sin cambios en la lógica principal)
        row += 1
        ws[f'A{row}'] = "MÉTRICAS DE EFICIENCIA"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        if analysis['start_time'] and analysis['end_time']:
            tiempo_calendario_min = (analysis['end_time'] - analysis['start_time']).total_seconds() / 60
            tiempo_productivo_total = analysis[
                'total_duration_min']  # Tiempo sumado de todas las tareas individuales

            # CORRECCIÓN Eficiencia: Tiempo productivo disponible = tiempo calendario * num trabajadores
            if num_trabajadores > 0:
                tiempo_total_disponible_trabajadores = tiempo_calendario_min * num_trabajadores
                eficiencia = (
                            tiempo_productivo_total / tiempo_total_disponible_trabajadores * 100) if tiempo_total_disponible_trabajadores > 0 else 0
            else:
                eficiencia = 0  # No se puede calcular si no hay trabajadores

            ws[f'A{row}'] = "Tiempo calendario total:"
            ws[f'B{row}'] = f"{tiempo_calendario_min:.1f} min ({tiempo_calendario_min / 60:.1f} horas)"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "Tiempo productivo total (suma tareas):"
            ws[f'B{row}'] = f"{tiempo_productivo_total:.1f} min ({tiempo_productivo_total / 60:.1f} horas)"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "Eficiencia global (basada en trabajadores):"
            ws[f'B{row}'] = f"{eficiencia:.1f}%"
            ws[f'A{row}'].font = Font(bold=True)
            if eficiencia >= 80:
                ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif eficiencia >= 60:
                ws[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row += 1

            # --- CORRECCIÓN: Usar unidades_totales_calculadas ---
            tiempo_promedio_por_unidad = tiempo_productivo_total / unidades_totales_calculadas if unidades_totales_calculadas > 0 else 0
            ws[f'A{row}'] = "Tiempo productivo promedio/unidad:"
            ws[f'B{row}'] = f"{tiempo_promedio_por_unidad:.1f} min"
            ws[f'A{row}'].font = Font(bold=True)
            # --- FIN CORRECCIÓN ---

        # Gráfico de distribución por departamento (Sin cambios)
        if analysis.get('departments'):
            row += 2  # Más espacio antes del gráfico
            ws[f'E{row}'] = "DISTRIBUCIÓN POR DEPARTAMENTO"
            ws[f'E{row}'].font = Font(size=12, bold=True)
            chart_pie = PieChart()  # Renombrado para evitar conflicto con bar_chart
            chart_pie.title = "Tiempo por Departamento (minutos)"
            chart_pie.style = 10
            chart_pie_row_start = row + 1  # Renombrado
            sorted_depts = sorted(analysis['departments'].items(), key=lambda x: x[1], reverse=True)
            current_chart_row = chart_pie_row_start  # Renombrado

            # Limpiar área de datos antiguos si existiera (buena práctica)
            max_dept_rows = len(sorted_depts)
            for r in range(chart_pie_row_start,
                            chart_pie_row_start + max_dept_rows + 5):  # Limpiar algunas filas extra
                ws[f'E{r}'] = None
                ws[f'F{r}'] = None

            # Escribir nuevos datos
            for dept, minutes in sorted_depts:
                ws[f'E{current_chart_row}'] = dept
                ws[f'F{current_chart_row}'] = round(minutes, 1)
                current_chart_row += 1

            chart_pie_row_end = current_chart_row - 1

            if chart_pie_row_end >= chart_pie_row_start:  # Asegurarse que hay datos
                data_ref_pie = Reference(ws, min_col=6, min_row=chart_pie_row_start, max_row=chart_pie_row_end)
                labels_ref_pie = Reference(ws, min_col=5, min_row=chart_pie_row_start, max_row=chart_pie_row_end)
                chart_pie.add_data(data_ref_pie,
                                    titles_from_data=False)  # No usar títulos de datos si los ponemos aparte
                chart_pie.set_categories(labels_ref_pie)
                chart_pie.dataLabels = DataLabelList()
                chart_pie.dataLabels.showVal = True
                chart_pie.dataLabels.showPercent = True

        # Gráfico de barras - Tiempo de producción por producto (Sin cambios)
        row += 15
        ws[f'E{row}'] = "TIEMPO DE PRODUCCIÓN POR PRODUCTO"
        ws[f'E{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'E{row}:G{row}')

        # Calcular tiempo total por producto (Sin cambios)
        product_times = defaultdict(float)
        if all_results:  # Verificar si hay datos
            for task in all_results:
                producto_codigo = task.get('Codigo Producto', 'N/A')
                producto_key = producto_codigo if producto_codigo != 'N/A' else 'Sin Código'
                duracion = float(task.get('Duracion (min)', 0))
                product_times[producto_key] += duracion

        sorted_products = sorted(product_times.items(), key=lambda x: x[1], reverse=True)

        if sorted_products:
            bar_chart_data_row = row + 1
            ws[f'E{bar_chart_data_row}'] = "Producto"
            ws[f'F{bar_chart_data_row}'] = "Tiempo (min)"
            ws[f'E{bar_chart_data_row}'].font = Font(bold=True, size=9)
            ws[f'F{bar_chart_data_row}'].font = Font(bold=True, size=9)
            bar_chart_data_row += 1
            bar_chart_start_data_row = bar_chart_data_row

            # Limpiar área de datos antiguos
            max_prod_rows = len(sorted_products[:10])
            for r in range(bar_chart_start_data_row, bar_chart_start_data_row + max_prod_rows + 5):
                ws[f'E{r}'] = None
                ws[f'F{r}'] = None

            # Escribir nuevos datos
            for producto, tiempo_min in sorted_products[:10]:
                ws[f'E{bar_chart_data_row}'] = producto
                ws[f'F{bar_chart_data_row}'] = round(tiempo_min, 1)
                ws[f'E{bar_chart_data_row}'].alignment = Alignment(horizontal="left")
                ws[f'F{bar_chart_data_row}'].alignment = Alignment(horizontal="center")
                bar_chart_data_row += 1

            bar_chart_end_data_row = bar_chart_data_row - 1

            if bar_chart_end_data_row >= bar_chart_start_data_row:  # Asegurarse que hay datos
                bar_chart = BarChart()
                bar_chart.type = "col"
                bar_chart.style = 10
                bar_chart.title = "Tiempo de Producción por Producto (Top 10)"
                bar_chart.y_axis.title = 'Minutos'
                bar_chart.x_axis.title = 'Productos'

                data_bar = Reference(ws, min_col=6, min_row=bar_chart_start_data_row - 1,
                                        max_row=bar_chart_end_data_row)
                cats_bar = Reference(ws, min_col=5, min_row=bar_chart_start_data_row,
                                        max_row=bar_chart_end_data_row)

                bar_chart.add_data(data_bar, titles_from_data=True)
                bar_chart.set_categories(cats_bar)
                bar_chart.height = 10
                bar_chart.width = 18
                bar_chart.dataLabels = DataLabelList()
                bar_chart.dataLabels.showVal = True

                if hasattr(self, 'logger'):  # Verificar si logger existe
                    self.logger.info(
                        f"✅ Gráfico de barras por producto creado con {len(sorted_products[:10])} productos")

        # Ajustar anchos de columna (Sin cambios)
        for col_num in range(1, 9):  # Columnas A a H
            column_letter = chr(64 + col_num)
            ws.column_dimensions[column_letter].width = 20

            self.logger.info(
                f"✅ Gráfico de barras por producto creado con {len(sorted_products[:10])} productos")

        for col in range(1, 9):
            column_letter = chr(64 + col)
            ws.column_dimensions[column_letter].width = 20

    def _crear_hoja_grupos_secuenciales(self, wb, all_results, production_flow):
        """
        Crea una hoja que muestra los grupos secuenciales de tareas.
        """
        ws = wb.create_sheet("Grupos Secuenciales")

        # Título
        ws['A1'] = "ANÁLISIS DE GRUPOS SECUENCIALES"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:G1')

        row = 3
        if not production_flow:
            ws[f'A{row}'] = "No hay información de flujo de producción disponible"
            ws[f'A{row}'].font = Font(italic=True)
            return

        # Encabezados
        headers = ["Grupo", "Trabajador Asignado", "Número de Tareas", "Duración Total (min)", "Primera Tarea",
                   "Última Tarea", "Estado"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Identificar grupos secuenciales
        grupo_num = 1
        for step in production_flow:
            if step.get('type') == 'sequential_group':
                trabajador = step.get('assigned_worker', 'N/A')
                tareas = step.get('tasks', [])
                num_tareas = len(tareas)

                # Calcular duración total del grupo
                duracion_total = sum(
                    float(task.get('task', {}).get('tiempo', 0))
                    for task in tareas
                )

                primera_tarea = tareas[0].get('task', {}).get('descripcion', 'N/A') if tareas else 'N/A'
                ultima_tarea = tareas[-1].get('task', {}).get('descripcion', 'N/A') if tareas else 'N/A'

                # Escribir datos
                ws[f'A{row}'] = f"Grupo {grupo_num}"
                ws[f'B{row}'] = trabajador
                ws[f'C{row}'] = num_tareas
                ws[f'D{row}'] = round(duracion_total, 2)
                ws[f'E{row}'] = primera_tarea[:50]  # Limitar longitud
                ws[f'F{row}'] = ultima_tarea[:50]
                ws[f'G{row}'] = "✓ Completado"

                ws[f'G{row}'].font = Font(color="008000")

                grupo_num += 1
                row += 1

        # Auto-ajustar columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 15

        self.logger.info(f"Hoja de grupos secuenciales creada con {grupo_num - 1} grupos")

    def _crear_hoja_cronograma(self, wb, all_results, hay_limite=False, total_original=None):
        """
        Crea una hoja con el cronograma detallado por UNIDAD INDIVIDUAL,
        ordenado cronológicamente, con separadores de día, y resaltando instancias paralelas.
        MODIFICADO: Añade columnas 'Instancia', 'Grupo Trabajo' y formato visual.
        """
        ws = wb.create_sheet("Cronograma Detallado")

        # Título principal
        ws['A1'] = "CRONOGRAMA DETALLADO POR UNIDAD Y ORDEN CRONOLÓGICO"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        # Extendemos el título para abarcar las NUEVAS columnas (ahora hasta M)
        ws.merge_cells('A1:M1') # <-- MODIFICADO

        if hay_limite and total_original:
            ws['A2'] = f"⚠️ Mostrando {len(all_results)} de {total_original} tareas (límite por optimización)"
            ws['A2'].font = Font(size=10, italic=True, color="FF0000")
            ws.merge_cells('A2:M2') # <-- MODIFICADO
            header_row = 4
        else:
            header_row = 3

        # Encabezados (con las nuevas columnas 'Instancia' y 'Grupo Trabajo')
        headers = [
            "#", "Inicio", "Fin", "Tarea",
            "Instancia",         # NUEVO
            "Grupo Trabajo",     # NUEVO
            "Trabajador(es)",    # Mantenido por compatibilidad o info general
            "Máquina", "Duración (min)", "Producto", "Unidad #", "Departamento", "Fab ID"
        ] # Total 13 columnas (A-M)

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row = header_row + 1
        last_date = None
        event_sequence = 0

        # --- INICIO: Lógica de Agrupación por Tarea y Formato ---
        # 1. Agrupar resultados por Nombre de Tarea
        datos_agrupados_por_tarea = defaultdict(list)
        for resultado in all_results:
            # Usar 'TareaDetalle' si existe, si no 'Tarea'
            tarea_nombre = resultado.get('TareaDetalle', resultado.get('Tarea', 'Tarea Desconocida'))
            datos_agrupados_por_tarea[tarea_nombre].append(resultado)

        # 2. Iterar sobre las tareas agrupadas
        for tarea_nombre, resultados_tarea in datos_agrupados_por_tarea.items():
            # Ordenar las unidades de ESTA tarea por inicio (importante para visualización)
            resultados_tarea.sort(key=lambda x: x.get('Inicio', datetime.min))

            # Detectar si hay instancias paralelas DENTRO de esta tarea
            instancias_en_tarea = set()
            for res in resultados_tarea:
                # Usar 'Instancia ID' si existe, si no, 'N/A'
                inst_id = res.get('Instancia ID', 'N/A')
                if inst_id != 'N/A':
                    instancias_en_tarea.add(inst_id)
            hay_paralelo_en_tarea = len(instancias_en_tarea) > 1

            self.logger.debug(f"Procesando Tarea '{tarea_nombre}': {len(resultados_tarea)} unidades. ¿Paralelo? {hay_paralelo_en_tarea}")

            # 3. Iterar sobre los resultados (unidades) de ESTA tarea
            for task in resultados_tarea:
                inicio = task.get('Inicio')
                if not isinstance(inicio, datetime):
                    self.logger.warning(f"Tarea '{task.get('Tarea')}' sin fecha de inicio válida. Omitiendo.")
                    continue

                current_date = inicio.date()

                # Insertar separador de día si la fecha cambia
                if current_date != last_date:
                    ws.merge_cells(f'A{current_row}:M{current_row}') # <-- MODIFICADO hasta M
                    cell_fecha = ws[f'A{current_row}']
                    cell_fecha.value = f"--- {current_date.strftime('%A, %d de %B de %Y')} ---"
                    cell_fecha.font = Font(size=11, bold=True, color="FFFFFF")
                    cell_fecha.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell_fecha.alignment = Alignment(horizontal="center")
                    current_row += 1
                    last_date = current_date

                # Escribir la fila de datos
                event_sequence += 1
                fin = task.get('Fin')

                # Extraer datos existentes
                inicio_str = inicio.strftime('%d/%m %H:%M') if inicio else 'N/A'
                fin_str = fin.strftime('%H:%M') if fin else 'N/A'
                nombre_tarea_raw = str(task.get('Tarea', 'Sin nombre')) # Nombre base
                # Mantenemos 'Trabajador Asignado' como info general si viene del resultado
                trabajadores_general = task.get('Trabajador Asignado', 'N/A')
                if isinstance(trabajadores_general, list):
                    trabajadores_general_str = ', '.join(trabajadores_general) if trabajadores_general else 'N/A'
                else:
                    trabajadores_general_str = str(trabajadores_general)

                maquina = str(task.get('nombre_maquina', 'N/A'))
                duracion = task.get('Duracion (min)', 0)
                product_code = task.get('Codigo Producto', '')
                product_desc = task.get('Descripcion Producto', '')
                producto = f"{product_code} - {product_desc}" if product_desc else product_code or 'N/A'
                numero_unidad = task.get('Numero Unidad', '?')
                departamento = str(task.get('Departamento', 'General'))
                fab_id = str(task.get('fabricacion_id', 'N/A'))

                # Extraer NUEVOS datos de instancia
                instancia_id_completo = task.get('Instancia ID', 'N/A')
                if instancia_id_completo != 'N/A':
                    instancia_id_corto = instancia_id_completo[:8]
                else:
                    instancia_id_corto = 'Principal'

                grupo_trabajo_list = task.get('Lista Trabajadores', []) # Esta SÍ es la lista específica de la instancia
                grupo_trabajo_str = ", ".join(grupo_trabajo_list) if grupo_trabajo_list else 'N/A'

                # Construir fila con nuevas columnas
                row_data = [
                    event_sequence, inicio_str, fin_str, nombre_tarea_raw,
                    instancia_id_corto,    # NUEVO (Col E)
                    grupo_trabajo_str,     # NUEVO (Col F)
                    trabajadores_general_str, # Mantenido (Col G)
                    maquina, duracion, producto, numero_unidad, departamento, fab_id
                ]

                # Escribir fila y aplicar formato básico
                for col_num, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    # Formatos (adaptados a nuevas columnas)
                    if col_num in [1, 11]: # #, Unidad #
                        cell.alignment = Alignment(horizontal="center")
                    elif col_num in [9]: # Duración
                        cell.number_format = '0.0'
                        cell.alignment = Alignment(horizontal="center")
                    elif col_num in [2, 3]: # Inicio, Fin
                        cell.alignment = Alignment(horizontal="center")
                    else: # Tarea, Instancia, Grupo, Trabajador(es), Máquina, Producto, Dept, FabID
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    # Bordes suaves (sin cambios)
                    cell.border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                                        top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

                # --- Aplicar Formato Condicional para Instancias Paralelas ---
                if hay_paralelo_en_tarea and instancia_id_completo != 'N/A':
                    # Determinar color basado en hash del ID
                    try:
                        color_idx = abs(hash(instancia_id_completo)) % 3 # Usamos hash para consistencia
                        colores_fondo = ['E8F4F8', 'F8E8F4', 'F4F8E8'] # Azul claro, Rosa claro, Verde claro
                        color_fondo = colores_fondo[color_idx]
                        instancia_fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type='solid')

                        # Aplicar relleno a toda la fila y borde izquierdo grueso
                        for col_idx_format in range(1, len(headers) + 1):
                            cell_format = ws.cell(row=current_row, column=col_idx_format)
                            cell_format.fill = instancia_fill
                            if col_idx_format == 1: # Solo en la primera columna
                                cell_format.border = Border(
                                    left=Side(style='thick', color='A0A0A0'), # Borde izquierdo grueso grisáceo
                                    right=Side(style='thin', color='E0E0E0'),
                                    top=Side(style='thin', color='E0E0E0'),
                                    bottom=Side(style='thin', color='E0E0E0')
                                )
                    except Exception as e_fmt:
                        self.logger.warning(f"Error aplicando formato de instancia: {e_fmt}")

                current_row += 1

            # --- Añadir separador visual entre tareas ---
            if resultados_tarea: # Solo si hubo filas para esta tarea
                last_row_of_task = current_row - 1
                for col_idx_sep in range(1, len(headers) + 1):
                    cell_sep = ws.cell(row=last_row_of_task, column=col_idx_sep)
                    # Añadir borde inferior medio-grueso
                    current_border = cell_sep.border
                    cell_sep.border = Border(
                        left=current_border.left, right=current_border.right,
                        top=current_border.top, bottom=Side(style='medium', color='B0B0B0') # Borde gris medio
                    )

        # --- FIN: Lógica de Agrupación y Formato ---

        # Ajustar anchos de columna (añadir nuevas y ajustar existentes)
        column_widths = {
            'A': 5,  # #
            'B': 18, # Inicio
            'C': 10, # Fin
            'D': 35, # Tarea
            'E': 12, # Instancia (NUEVO)
            'F': 25, # Grupo Trabajo (NUEVO)
            'G': 25, # Trabajador(es)
            'H': 15, # Máquina
            'I': 10, # Duración (min) - Cambiado de G a I
            'J': 30, # Producto - Cambiado de H a J
            'K': 10, # Unidad # - Cambiado de I a K
            'L': 15, # Departamento - Cambiado de J a L
            'M': 12  # Fab ID - Cambiado de K a M
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Congelar primera fila (encabezados)
        ws.freeze_panes = f"A{header_row + 1}"

        # Aplicar filtros automáticos
        ws.auto_filter.ref = ws.dimensions

        self.logger.info(f"✅ Hoja de cronograma detallado creada/actualizada con {len(all_results)} eventos, formato de instancia y separadores.")

    def _crear_hoja_cuellos_botella(self, wb, audit_log):
        """
        Crea una hoja con análisis ULTRA DETALLADO de cuellos de botella y tiempos muertos.
        Incluye análisis cruzado con información de tareas, trabajadores y tiempos específicos.
        """
        ws = wb.create_sheet("Cuellos de Botella")

        # Título
        ws['A1'] = "ANÁLISIS ULTRA DETALLADO DE CUELLOS DE BOTELLA Y TIEMPOS MUERTOS"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws.merge_cells('A1:N1')

        # Obtener all_results para hacer cruce de información
        all_results = self.workbook._datos_informe.get("data", []) if hasattr(self.workbook, '_datos_informe') else []

        row = 3

        # ========================================================================
        # ANÁLISIS DE TIEMPOS INACTIVOS CON INFORMACIÓN COMPLETA
        # ========================================================================

        ws[f'A{row}'] = "⏸️ ANÁLISIS DETALLADO DE TIEMPOS INACTIVOS"
        ws[f'A{row}'].font = Font(size=13, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        ws.merge_cells(f'A{row}:N{row}')
        row += 2

        # Filtrar eventos de tiempo inactivo
        tiempos_inactivos_detallados = []

        for event in audit_log:
            if isinstance(event, CalculationDecision):
                if 'TIEMPO_INACTIVO' in event.decision_type.upper() or 'INACTIV' in event.decision_type.upper():
                    details = event.details if hasattr(event, 'details') and event.details else {}

                    # Información básica del evento
                    timestamp = event.timestamp
                    trabajador_inactivo = details.get('trabajador', 'N/A')
                    tarea_completada = details.get('tarea_actual', event.task_name)
                    proxima_tarea_texto = details.get('proxima_tarea', 'N/A')
                    tarea_bloqueante = details.get('esperando_a', 'N/A')
                    duracion_espera_min = details.get('wait_time', 0) or details.get('wait_minutes', 0)

                    # Extraer número de unidad de la tarea completada y próxima
                    import re

                    # De tarea_completada, intentar extraer el número de unidad
                    unidad_completada = 'N/A'
                    match_completada = re.search(r'U(\d+)', str(tarea_completada))
                    if match_completada:
                        unidad_completada = match_completada.group(1)

                    # De proxima_tarea, extraer número de unidad
                    unidad_proxima = 'N/A'
                    match_proxima = re.search(r'U(\d+)', str(proxima_tarea_texto))
                    if match_proxima:
                        unidad_proxima = match_proxima.group(1)

                    # Calcular cuándo podrá comenzar (timestamp + duracion)
                    hora_finalizacion_inactividad = timestamp + timedelta(minutes=duracion_espera_min)

                    # CRUCE DE INFORMACIÓN: Buscar quién está trabajando en la tarea bloqueante
                    trabajadores_bloqueantes = []
                    unidad_bloqueante = 'N/A'
                    hora_fin_tarea_bloqueante = None

                    if all_results:
                        # Buscar en all_results la tarea bloqueante que se completará después del timestamp actual
                        for task in all_results:
                            tarea_nombre = task.get('Tarea', '')
                            fin_tarea = task.get('Fin')

                            # Si la tarea coincide con la bloqueante y termina después del evento de inactividad
                            if tarea_bloqueante in tarea_nombre and fin_tarea and fin_tarea > timestamp:
                                # Esta podría ser la tarea que está bloqueando
                                trabajadores_en_tarea = task.get('Trabajador Asignado', [])
                                if isinstance(trabajadores_en_tarea, str):
                                    trabajadores_en_tarea = [t.strip() for t in trabajadores_en_tarea.split(',')]

                                # Obtener número de unidad
                                numero_unidad_task = task.get('Numero Unidad', '?')

                                # Si encontramos una tarea que termina aproximadamente cuando debería terminar la espera
                                tiempo_diferencia = abs(
                                    (fin_tarea - timestamp).total_seconds() / 60 - duracion_espera_min)

                                if tiempo_diferencia < 30:  # Tolerancia de 30 minutos
                                    trabajadores_bloqueantes = trabajadores_en_tarea
                                    unidad_bloqueante = numero_unidad_task
                                    hora_fin_tarea_bloqueante = fin_tarea
                                    break

                    # Si no encontramos mediante cruce, intentar extraer de los eventos futuros
                    if not trabajadores_bloqueantes:
                        trabajadores_bloqueantes = ['Información no disponible']

                    trabajadores_str = ', '.join(trabajadores_bloqueantes) if isinstance(trabajadores_bloqueantes,
                                                                                         list) else str(
                        trabajadores_bloqueantes)

                    tiempos_inactivos_detallados.append({
                        'timestamp': timestamp,
                        'trabajador_inactivo': trabajador_inactivo,
                        'tarea_completada': tarea_completada,
                        'unidad_completada': unidad_completada,
                        'proxima_tarea': proxima_tarea_texto,
                        'unidad_proxima': unidad_proxima,
                        'tarea_bloqueante': tarea_bloqueante,
                        'unidad_bloqueante': unidad_bloqueante,
                        'trabajadores_bloqueantes': trabajadores_str,
                        'duracion_espera_min': duracion_espera_min,
                        'duracion_espera_horas': duracion_espera_min / 60,
                        'hora_fin_espera': hora_finalizacion_inactividad,
                        'hora_fin_tarea_bloqueante': hora_fin_tarea_bloqueante,
                        'reason_completo': event.reason if hasattr(event, 'reason') else ''
                    })

        if not tiempos_inactivos_detallados:
            ws[f'A{row}'] = "✅ No se detectaron tiempos inactivos en esta simulación"
            ws[f'A{row}'].font = Font(size=11, bold=True, color="008000")
            ws.merge_cells(f'A{row}:N{row}')
            row += 3
        else:
            # Resumen estadístico
            total_tiempo_inactivo = sum(ti['duracion_espera_min'] for ti in tiempos_inactivos_detallados)
            promedio_inactividad = total_tiempo_inactivo / len(tiempos_inactivos_detallados)
            max_inactividad = max(ti['duracion_espera_min'] for ti in tiempos_inactivos_detallados)
            eventos_criticos = sum(1 for ti in tiempos_inactivos_detallados if ti['duracion_espera_min'] > 60)

            ws[f'A{row}'] = "📊 RESUMEN"
            ws[f'A{row}'].font = Font(size=11, bold=True, underline='single')
            row += 1

            resumen = [
                ("Total eventos:", f"{len(tiempos_inactivos_detallados)} eventos"),
                ("Tiempo total perdido:",
                 f"{total_tiempo_inactivo:.0f} min ({total_tiempo_inactivo / 60:.1f} h / {total_tiempo_inactivo / 480:.1f} jornadas)"),
                ("Promedio por evento:", f"{promedio_inactividad:.0f} min"),
                ("Mayor inactividad:", f"{max_inactividad:.0f} min ({max_inactividad / 60:.1f} horas)"),
                ("Eventos críticos (>60 min):", f"{eventos_criticos} eventos")
            ]

            for label, value in resumen:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                if "críticos" in label and eventos_criticos > 0:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'B{row}'].font = Font(bold=True, color="C00000")
                row += 1

            row += 2

            # Tabla ultra detallada
            ws[f'A{row}'] = "📋 DETALLE COMPLETO POR EVENTO"
            ws[f'A{row}'].font = Font(size=11, bold=True, underline='single')
            ws.merge_cells(f'A{row}:N{row}')
            row += 1

            # Encabezados expandidos
            headers = [
                "Fecha/Hora\nInicio Espera",
                "Trabajador\nInactivo",
                "Tarea\nCompletada",
                "Unidad\nCompletada",
                "Siguiente\nTarea",
                "Unidad\nSiguiente",
                "Esperando\nTarea",
                "Unidad\nBloqueante",
                "Trabajador(es)\nen Tarea Bloqueante",
                "Termina\nTarea Bloqueante",
                "Podrá Comenzar\nSiguiente Tarea",
                "Espera\n(min)",
                "Espera\n(horas)",
                "Severidad"
            ]

            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            row += 1

            # Ordenar por duración descendente
            tiempos_inactivos_detallados.sort(key=lambda x: x['duracion_espera_min'], reverse=True)

            # Datos
            for ti in tiempos_inactivos_detallados:
                # Preparar datos
                timestamp_str = ti['timestamp'].strftime('%d/%m/%Y\n%H:%M') if isinstance(ti['timestamp'],
                                                                                          datetime) else str(
                    ti['timestamp'])
                fin_bloqueante_str = ti['hora_fin_tarea_bloqueante'].strftime('%d/%m/%Y\n%H:%M') if ti[
                    'hora_fin_tarea_bloqueante'] else 'N/A'
                podra_comenzar_str = ti['hora_fin_espera'].strftime('%d/%m/%Y\n%H:%M') if ti[
                    'hora_fin_espera'] else 'N/A'

                # Determinar severidad
                if ti['duracion_espera_min'] > 240:  # >4 horas
                    severidad = "🔴 CRÍTICO"
                    fill_color = "C00000"
                    font_color = "FFFFFF"
                elif ti['duracion_espera_min'] > 60:  # >1 hora
                    severidad = "🟠 ALTO"
                    fill_color = "FFC7CE"
                    font_color = "9C0006"
                elif ti['duracion_espera_min'] > 30:
                    severidad = "🟡 MEDIO"
                    fill_color = "FFEB9C"
                    font_color = "9C6500"
                else:
                    severidad = "🟢 BAJO"
                    fill_color = "C6EFCE"
                    font_color = "006100"

                # Escribir fila
                datos_fila = [
                    timestamp_str,
                    ti['trabajador_inactivo'],
                    ti['tarea_completada'],
                    ti['unidad_completada'],
                    ti['proxima_tarea'],
                    ti['unidad_proxima'],
                    ti['tarea_bloqueante'],
                    ti['unidad_bloqueante'],
                    ti['trabajadores_bloqueantes'],
                    fin_bloqueante_str,
                    podra_comenzar_str,
                    round(ti['duracion_espera_min'], 0),
                    round(ti['duracion_espera_horas'], 1),
                    severidad
                ]

                for col_num, valor in enumerate(datos_fila, start=1):
                    cell = ws.cell(row=row, column=col_num)
                    cell.value = valor
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(size=9, color=font_color if col_num == 14 else "000000")

                    if col_num == 14:  # Columna de severidad
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.font = Font(size=9, bold=True, color=font_color)

                row += 1

            row += 2

            # Añadir explicación narrativa para el caso más crítico
            evento_mas_critico = max(tiempos_inactivos_detallados, key=lambda x: x['duracion_espera_min'])

            ws[f'A{row}'] = "💡 ANÁLISIS DEL EVENTO MÁS CRÍTICO"
            ws[f'A{row}'].font = Font(size=11, bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            ws.merge_cells(f'A{row}:N{row}')
            row += 1

            # Construir explicación narrativa
            explicacion = (
                f"El {evento_mas_critico['timestamp'].strftime('%d/%m/%Y a las %H:%M')}, "
                f"el trabajador {evento_mas_critico['trabajador_inactivo']} finalizó la tarea "
                f"'{evento_mas_critico['tarea_completada']}' (Unidad {evento_mas_critico['unidad_completada']}) "
                f"y debe esperar {evento_mas_critico['duracion_espera_min']:.0f} minutos "
                f"({evento_mas_critico['duracion_espera_horas']:.1f} horas) para comenzar su siguiente tarea "
                f"'{evento_mas_critico['proxima_tarea']}' (Unidad {evento_mas_critico['unidad_proxima']}). "
                f"\n\nEsta espera se debe a que necesita que se complete primero la tarea "
                f"'{evento_mas_critico['tarea_bloqueante']}' (Unidad {evento_mas_critico['unidad_bloqueante']}), "
                f"que está siendo ejecutada por: {evento_mas_critico['trabajadores_bloqueantes']}. "
            )

            if evento_mas_critico['hora_fin_tarea_bloqueante']:
                explicacion += (
                    f"\n\nLa tarea bloqueante finalizará el "
                    f"{evento_mas_critico['hora_fin_tarea_bloqueante'].strftime('%d/%m/%Y a las %H:%M')}, "
                    f"momento en el cual {evento_mas_critico['trabajador_inactivo']} podrá "
                    f"comenzar su trabajo en '{evento_mas_critico['proxima_tarea']}'."
                )

            ws[f'A{row}'] = explicacion
            ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws[f'A{row}'].font = Font(size=10)
            ws.merge_cells(f'A{row}:N{row}')
            ws.row_dimensions[row].height = 80
            row += 3

            # Gráfico
            ws[f'A{row}'] = "📊 GRÁFICO: TIEMPOS INACTIVOS POR TRABAJADOR"
            ws[f'A{row}'].font = Font(size=11, bold=True, underline='single')
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            # Agrupar por trabajador
            trabajador_total = defaultdict(float)
            for ti in tiempos_inactivos_detallados:
                trabajador_total[ti['trabajador_inactivo']] += ti['duracion_espera_min']

            sorted_trabajadores = sorted(trabajador_total.items(), key=lambda x: x[1], reverse=True)

            chart_start = row
            ws[f'A{chart_start}'] = "Trabajador"
            ws[f'B{chart_start}'] = "Tiempo Inactivo (min)"
            ws[f'A{chart_start}'].font = Font(bold=True)
            ws[f'B{chart_start}'].font = Font(bold=True)
            chart_start += 1

            for trabajador, tiempo in sorted_trabajadores:
                ws[f'A{chart_start}'] = trabajador
                ws[f'B{chart_start}'] = round(tiempo, 1)
                chart_start += 1

            chart_end = chart_start - 1

            if chart_end > row + 1:
                chart_inactivos = BarChart()
                chart_inactivos.type = "col"
                chart_inactivos.style = 12
                chart_inactivos.title = "Tiempo Total de Inactividad por Trabajador"
                chart_inactivos.y_axis.title = 'Minutos Inactivos'
                chart_inactivos.x_axis.title = 'Trabajadores'
                chart_inactivos.height = 10
                chart_inactivos.width = 16

                data_ref = Reference(ws, min_col=2, min_row=row + 1, max_row=chart_end)
                cats_ref = Reference(ws, min_col=1, min_row=row + 2, max_row=chart_end)

                chart_inactivos.add_data(data_ref, titles_from_data=True)
                chart_inactivos.set_categories(cats_ref)

                ws.add_chart(chart_inactivos, f'D{row}')

        # Ajustar anchos de columna
        column_widths = {
            'A': 13, 'B': 15, 'C': 18, 'D': 10, 'E': 18, 'F': 10,
            'G': 18, 'H': 10, 'I': 25, 'J': 13, 'K': 13,
            'L': 10, 'M': 10, 'N': 12
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        self.logger.info(
            f"✅ Hoja ultra detallada de cuellos de botella creada con {len(tiempos_inactivos_detallados)} eventos")

    def _generar_descripcion_cuello_botella(self, event):
        """
        Genera una descripción detallada y legible del cuello de botella.
        """
        descripcion_partes = []

        # Información básica del evento
        if event.get('reason'):
            descripcion_partes.append(event['reason'])

        # Añadir información de tiempo
        if event['impact_min'] > 0:
            horas = event['impact_hours']
            jornadas = event['impact_jornadas']

            tiempo_str = f"Tiempo de espera: {event['impact_min']:.1f} minutos"
            if horas >= 1:
                tiempo_str += f" ({horas:.1f} horas)"
            if jornadas >= 0.5:
                tiempo_str += f" (aproximadamente {jornadas:.1f} jornadas laborales)"

            descripcion_partes.append(tiempo_str)

        # Información del recurso bloqueante
        if event['resource'] != 'N/A':
            descripcion_partes.append(f"Recurso bloqueante: {event['resource']}")

        # Información del producto
        if event['producto_esperado'] != 'N/A':
            descripcion_partes.append(f"Esperando material de: {event['producto_esperado']}")

        return " | ".join(descripcion_partes) if descripcion_partes else "Sin información adicional"

    def guardar_reporte(self, output_path: str) -> bool:
        if not self.workbook:
            self.logger.error("No hay un workbook para guardar. Ejecute generar_reporte() primero.")
            return False
        try:
            self.workbook.save(output_path)
            self.logger.info(f"Reporte Excel guardado en: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Error al guardar el archivo Excel: {e}", exc_info=True)
            return False

    def _crear_hoja_trabajo_paralelo(self, wb, all_results):
        """
        Crea una hoja dedicada al análisis de trabajo paralelo,
        agrupando por instancia.
        (Basado en la Propuesta 6.2 del plan)
        """
        ws = wb.create_sheet("Trabajo Paralelo")

        # Título
        ws['A1'] = "ANÁLISIS DE TRABAJO PARALELO POR INSTANCIA"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:H1')

        # Filtrar solo resultados con instancias paralelas válidas
        resultados_paralelos = []
        for res in all_results:
            inst_id = res.get('Instancia ID')
            if inst_id and inst_id != 'N/A' and inst_id != 'Principal':
                resultados_paralelos.append(res)

        if not resultados_paralelos:
            ws['A3'] = "No se detectó trabajo paralelo (instancias múltiples) en esta simulación."
            ws['A3'].font = Font(italic=True)
            self.logger.info("Hoja 'Trabajo Paralelo' creada (sin datos).")
            return

        # Agrupar por instancia ID
        instancias = defaultdict(list)
        for res in resultados_paralelos:
            inst_id = res['Instancia ID']
            instancias[inst_id].append(res)

        # Encabezados de la tabla
        headers = [
            "Tarea",
            "Instancia ID (8 dígitos)",
            "Trabajadores en Instancia",
            "Unidad Inicial",
            "Unidad Final",
            "Inicio",
            "Fin",
            "Duración Total (min)"
        ]

        current_row = 3
        ws.append(headers)

        # Aplicar formato a encabezados
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # Añadir datos agregados por instancia
        # Ordenar por Tarea y luego por ID de instancia
        for inst_id, registros in sorted(instancias.items(),
                                            key=lambda item: (item[1][0].get('Tarea', ''), item[0])):

            # Ordenar registros de esta instancia por inicio
            registros.sort(key=lambda x: x.get('Inicio', datetime.min))

            # Calcular agregados
            tarea = registros[0].get('Tarea', 'N/A')
            trabajadores = registros[0].get('Lista Trabajadores', [])
            unidad_inicial = min(r.get('Numero Unidad', 0) for r in registros)
            unidad_final = max(r.get('Numero Unidad', 0) for r in registros)
            inicio = min(r.get('Inicio') for r in registros if r.get('Inicio'))
            fin = max(r.get('Fin') for r in registros if r.get('Fin'))
            duracion_total = sum(r.get('Duracion (min)', 0) for r in registros)

            row_data = [
                tarea,
                inst_id[:8],  # ID Abreviado
                ", ".join(trabajadores),
                unidad_inicial,
                unidad_final,
                inicio.strftime('%d/%m/%Y %H:%M') if inicio else 'N/A',
                fin.strftime('%d/%m/%Y %H:%M') if fin else 'N/A',
                round(duracion_total, 2)
            ]

            ws.append(row_data)

            # Formato alternado (filas de datos)
            if (current_row - 3) % 2 == 0:  # (fila_actual - fila_header)
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=current_row, column=col_idx).fill = fill

            current_row += 1

        # Ajustar anchos
        column_widths = [30, 18, 30, 12, 12, 20, 20, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Congelar primera fila (encabezados)
        ws.freeze_panes = "A4"

        self.logger.info(f"Hoja 'Trabajo Paralelo' creada con {len(instancias)} instancias detalladas.")

class ReporteHistorialFabricacion(IReporteEstrategia):

    def __init__(self, model, schedule_config=None):
        self.model = model
        self.schedule_config = schedule_config
    """
    Estrategia para generar un informe PDF de optimización, incluyendo
    resumen ejecutivo, diagnóstico de cuellos de botella y log detallado.
    """

    def generar_reporte(self, datos_informe, output_path) -> bool:
        self.logger = logging.getLogger("EvolucionTiemposApp")
        self.logger.info(f"Generando informe PDF evolucionado en: {output_path}")

        try:
            # Extracción de datos
            meta_data = datos_informe.get("meta_data", {})
            results = datos_informe.get("planificacion", [])
            audit = datos_informe.get("audit", [])
            workers_needed = datos_informe.get("flexible_workers_needed", 0)
            production_flow = datos_informe.get("production_flow", [])

            if not results:
                self.logger.error("No hay datos de planificación para el PDF.")
                return False

            doc = SimpleDocTemplate(output_path, pagesize=landscape(A4), topMargin=inch / 2, bottomMargin=inch / 2)
            styles = getSampleStyleSheet()
            story = []

            # 1. Portada y Resumen Ejecutivo
            self._add_executive_summary(story, datos_informe.get("meta_data", {}),
                                        datos_informe.get("flexible_workers_needed", 0), results, styles)
            story.append(PageBreak())

            # 2. Cronograma Visual (Gantt)
            story.append(Paragraph("Cronograma Visual de Planificación (Gantt)", styles['h2']))
            self._add_gantt_chart_to_pdf(story, results, styles)
            story.append(Spacer(1, 0.25 * inch))

            # 3. Análisis de Recursos y Diagnóstico
            story.append(Paragraph("Análisis y Diagnóstico de Recursos", styles['h2']))

            # Se realiza UNA SOLA LLAMADA con todos los argumentos correctos
            self._add_resource_analysis_to_pdf(story, results, self.model, production_flow, styles)

            story.append(Spacer(1, 0.25 * inch))
            self._add_sequential_group_diagnostics(story, audit, styles)
            story.append(Spacer(1, 0.25 * inch))
            self._add_diagnostics(story, audit, styles)
            story.append(PageBreak())

            # 4. Log de Auditoría Detallado
            story.append(Paragraph("Log de Decisiones Detallado", styles['h2']))
            self._add_audit_log_table(story, audit, styles)

            doc.build(story)
            self.logger.info("Informe PDF evolucionado generado con éxito.")
            return True
        except Exception as e:
            self.logger.critical(f"Error al generar el informe PDF evolucionado: {e}", exc_info=True)
            return False

    def _add_executive_summary(self, story, meta_data, workers_needed, results, styles):
        story.append(Paragraph(f"Informe de Planificación de Lote: {meta_data.get('code', 'N/A')}", styles['h1']))
        story.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.5 * inch))

        story.append(Paragraph("Resumen Ejecutivo de la Planificación", styles['h2']))

        # --- CORRECCIÓN EN EL TEXTO DEL RESUMEN ---
        summary_text = (
            f"Para la configuración y plazos definidos, el sistema ha determinado que se requiere un total de "
            f"<font size=12 color='blue'><b>{workers_needed} trabajador(es) flexible(s)</b></font> "
            f"adicional(es) al personal especialista existente para cumplir los objetivos."
        )
        story.append(Paragraph(summary_text, styles['BodyText']))
        story.append(Spacer(1, 0.3 * inch))

        start_time = min(r['Inicio'] for r in results)
        end_time = max(r['Fin'] for r in results)

        total_workdays = self.time_calculator.count_workdays(start_time, end_time)

        data = [
            [Paragraph('<b>Fecha de Inicio Estimada:</b>', styles['Normal']), start_time.strftime('%d/%m/%Y %H:%M')],
            [Paragraph('<b>Fecha de Fin Estimada:</b>', styles['Normal']), end_time.strftime('%d/%m/%Y %H:%M')],
            [Paragraph('<b>Duración Total (Jornadas):</b>', styles['Normal']), f"{total_workdays:.2f}"],
            [Paragraph('<b>Número de Tareas Totales:</b>', styles['Normal']), str(len(results))],
        ]
        table = Table(data, colWidths=[2.5 * inch, 5 * inch])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(table)

    def _add_gantt_chart_to_pdf(self, story, results, styles):
        """Crea un Gantt simplificado usando una tabla de ReportLab."""
        if not results: return

        start_date = min(r['Inicio'] for r in results).date()
        end_date = max(r['Fin'] for r in results).date()
        total_days = (end_date - start_date).days + 1

        # --- Preparación de la tabla ---
        # Cabecera con los días
        header = ['<b>Tarea</b>'] + [(start_date + timedelta(days=i)).strftime('%d/%m') for i in range(total_days)]
        table_data = [header]

        # Filas con los datos de cada tarea
        for task in results:
            task_start_date = task['Inicio'].date()
            task_end_date = task['Fin'].date()

            row = [Paragraph(task['Tarea'], styles['Normal'])]
            for i in range(total_days):
                current_cal_date = start_date + timedelta(days=i)
                if task_start_date <= current_cal_date <= task_end_date:
                    row.append("")  # Dejamos la celda vacía, la colorearemos
                else:
                    row.append("")
            table_data.append(row)

        # --- Creación y Estilizado de la Tabla ---
        # Anchos de columna: más grande para el nombre, pequeñas para los días
        col_widths = [2.5 * inch] + [0.2 * inch] * total_days

        # Cortafuegos para evitar tablas excesivamente anchas
        if len(col_widths) > 50:
            col_widths = [2.5 * inch] + [0.15 * inch] * total_days

        table = Table(table_data, colWidths=col_widths)

        # Estilos de la tabla
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Cabecera de fechas
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ])

        # Colorear las barras de las tareas
        for row_idx, task in enumerate(results, 1):
            start_offset = (task['Inicio'].date() - start_date).days
            end_offset = (task['Fin'].date() - start_date).days

            # Asignar un color por departamento
            dept_color = {
                'Mecánica': colors.HexColor('#3498db'),
                'Electrónica': colors.HexColor('#2ecc71'),
                'Montaje': colors.HexColor('#f1c40f')
            }.get(task.get('Departamento'), colors.grey)

            style.add('BACKGROUND', (start_offset + 1, row_idx), (end_offset + 1, row_idx), dept_color)

        table.setStyle(style)
        story.append(table)
        # --- INICIO: Añadir nota sobre trabajo paralelo (Opción Simple) ---
        story.append(Spacer(1, 0.1 * inch))
        nota_paralelo = Paragraph(
            "<i>Nota: Las tareas con múltiples grupos de trabajadores simultáneos (trabajo paralelo) "
            "se detallan en la sección 'Análisis y Diagnóstico de Recursos'.</i>",
            styles['Normal']
        )
        story.append(nota_paralelo)


    def _add_resource_analysis_to_pdf(self, story, results, model, production_flow, styles):
        """Analiza la asignación de trabajadores a tareas y su nivel de habilidad."""
        story.append(Paragraph("Análisis de Asignación de Habilidades", styles['h3']))

        all_workers = {w.nombre_completo: w.tipo_trabajador for w in model.worker_repo.get_all_workers(True)}  # {nombre: nivel}

        table_data = [
            ['<b>Tarea</b>', '<b>Trabajador(es) Asignado(s)</b>', '<b>Nivel Requerido</b>', '<b>Nivel Asignado</b>',
                '<b>Diagnóstico</b>']]

        inefficiencies = 0

        for task_result in results:
            task_name = task_result['Tarea']
            assigned_workers = task_result['Trabajador Asignado']


            original_task_data = None
            for step in production_flow:
                if step.get('type') == 'sequential_group':
                    # Si es un grupo, buscar dentro de sus tareas
                    found = next((t_wrapper['task'] for t_wrapper in step.get('tasks', []) if
                                    t_wrapper.get('task', {}).get('name') == task_name), None)
                    if found:
                        original_task_data = found
                        break
                else:
                    # Si es una tarea individual
                    if step.get('task', {}).get('name') == task_name:
                        original_task_data = step['task']
                        break

            if not original_task_data:
                continue

            required_skill = original_task_data.get('required_skill_level', 1)

            for worker_name in assigned_workers:
                worker_skill = all_workers.get(worker_name, 0)
                diagnostico = "Óptimo"
                if worker_skill > required_skill + 1:
                    diagnostico = "Sobrecalificado"
                    inefficiencies += 1
                elif worker_skill < required_skill:
                    diagnostico = "Subcalificado"
                    inefficiencies += 1

                table_data.append([
                    Paragraph(task_name, styles['BodyText']),
                    worker_name,
                    str(required_skill),
                    str(worker_skill),
                    diagnostico
                ])

        if len(table_data) == 1:
            story.append(Paragraph("No hay datos de asignación para analizar.", styles['BodyText']))
            return

        table = Table(table_data, colWidths=[3 * inch, 2 * inch, 1 * inch, 1 * inch, 1.5 * inch])

        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])

        # Colorear filas con diagnóstico no óptimo
        for i, row in enumerate(table_data):
            if row[-1] == "Sobrecalificado":
                style.add('BACKGROUND', (0, i), (-1, i), colors.lightgoldenrodyellow)
            elif row[-1] == "Subcalificado":
                style.add('BACKGROUND', (0, i), (-1, i), colors.lightpink)

        table.setStyle(style)
        story.append(table)
        story.append(Spacer(1, 0.1 * inch))
        story.append(Paragraph(f"Se detectaron <b>{inefficiencies}</b> asignaciones potencialmente ineficientes.",
                                styles['BodyText']))
        # --- INICIO: NUEVA SECCIÓN - Trabajo Paralelo Detectado ---
        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("Detección de Trabajo Paralelo", styles['h3']))

        tareas_con_paralelo = defaultdict(set)
        # Recopilar todas las instancias por tarea
        for res in results:
            # Asegurarse de que 'Instancia ID' existe y no es 'N/A' o None
            inst_id = res.get('Instancia ID')
            if inst_id and inst_id != 'N/A':
                # Usar 'Tarea' como clave
                tarea_nombre = res.get('Tarea', 'Desconocida')
                tareas_con_paralelo[tarea_nombre].add(inst_id)

        # Filtrar solo las tareas que tienen MÁS de una instancia
        tareas_paralelas_info = {
            tarea: instancias
            for tarea, instancias in tareas_con_paralelo.items()
            if len(instancias) > 1
        }

        if not tareas_paralelas_info:
            story.append(Paragraph(
                "No se detectó trabajo paralelo significativo (múltiples instancias simultáneas en la misma tarea).",
                styles['BodyText']
            ))
        else:
            story.append(Paragraph(
                f"Se detectaron <b>{len(tareas_paralelas_info)} tarea(s)</b> donde diferentes grupos de trabajadores "
                f"operaron en paralelo sobre distintas unidades:",
                styles['BodyText']
            ))
            story.append(Spacer(1, 0.1 * inch))

            # Crear tabla para mostrar detalles
            data_paralelo = [[
                Paragraph('<b>Tarea</b>', styles['Normal']),
                Paragraph('<b>Nº Instancias</b>', styles['Normal']),
                Paragraph('<b>IDs de Instancia (abreviados)</b>', styles['Normal'])
            ]]

            # Ordenar por nombre de tarea para consistencia
            for tarea, instancias_set in sorted(tareas_paralelas_info.items()):
                # Convertir set a lista y abreviar IDs
                ids_abreviados = sorted([inst_id[:8] for inst_id in instancias_set])  # Ordenar IDs también
                data_paralelo.append([
                    Paragraph(tarea, styles['BodyText']),
                    str(len(instancias_set)),
                    Paragraph(", ".join(ids_abreviados), styles['BodyText'])  # Usar Paragraph para posible wrap
                ])

            # Definir anchos de columna (ajustar según necesidad)
            table_paralelo = Table(data_paralelo, colWidths=[3 * inch, 1.2 * inch, 3.3 * inch])

            # Aplicar estilo a la tabla
            table_paralelo.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(table_paralelo)

    def _add_diagnostics(self, story, audit, styles):
        """
        Analiza el log para identificar cuellos de botella y tiempos de inactividad.
        VERSIÓN MEJORADA: Incluye el análisis de TIEMPO_INACTIVO.
        """
        # --- SECCIÓN EXISTENTE: Cuellos de botella por espera de recursos ---
        resource_warnings = [d for d in audit if d.decision_type == 'ESPERA POR RECURSO']
        story.append(Paragraph("Cuellos de Botella de Recursos (Máquinas/Trabajadores)", styles['h3']))

        if not resource_warnings:
            story.append(Paragraph("No se han detectado esperas significativas por recursos.", styles['BodyText']))
        else:
            total_wait_time = sum(d.details.get('wait_minutes', 0) for d in resource_warnings)
            story.append(Paragraph(f"Tiempo total de espera por recursos: <b>{total_wait_time:.1f} minutos</b>.",
                                   styles['BodyText']))
            for decision in sorted(resource_warnings, key=lambda d: d.details.get('wait_minutes', 0), reverse=True)[:3]:
                details = decision.details
                texto = (
                    f" • <b>{decision.task_name}:</b> Espera de <b>{details.get('wait_minutes', 0):.1f} min</b> "
                    f"por <i>'{details.get('resource', 'N/A')}'</i>."
                )
                story.append(Paragraph(texto, styles['Bullet']))

        story.append(Spacer(1, 0.2 * inch))

        # --- INICIO DE LA NUEVA SECCIÓN ---
        # Nueva sección para analizar los Tiempos de Inactividad por Dependencias
        idle_events = [d for d in audit if d.decision_type == "TIEMPO_INACTIVO"]
        story.append(Paragraph("Tiempos de Inactividad por Dependencias", styles['h3']))

        if not idle_events:
            story.append(Paragraph("No se han detectado parones en el flujo de trabajo por dependencias entre tareas.",
                                   styles['BodyText']))
        else:
            total_idle_time = sum((d.end_date - d.start_date).total_seconds() / 60 for d in idle_events)
            story.append(Paragraph(
                f"El flujo de producción se detuvo por un total de <b>{total_idle_time:.1f} minutos</b> esperando que se completaran tareas previas.",
                styles['BodyText']))
            story.append(Spacer(1, 0.1 * inch))

            # Creamos una tabla con los 5 parones más largos
            table_data = [['<b>Inicio de la Inactividad</b>', '<b>Fin de la Inactividad</b>', '<b>Duración (min)</b>']]

            # Ordenamos los eventos por duración para mostrar los más significativos
            sorted_idle_events = sorted(idle_events, key=lambda d: (d.end_date - d.start_date), reverse=True)

            for event in sorted_idle_events[:5]:  # Mostramos hasta 5
                duration = (event.end_date - event.start_date).total_seconds() / 60
                table_data.append([
                    event.start_date.strftime('%d/%m %H:%M'),
                    event.end_date.strftime('%d/%m %H:%M'),
                    f"{duration:.1f}"
                ])

            table = Table(table_data, colWidths=[2.5 * inch, 2.5 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#BF2A2A')),  # Mismo rojo que en Excel
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            story.append(table)
        # --- FIN DE LA NUEVA SECCIÓN ---

        story.append(Spacer(1, 0.2 * inch))

        # --- SECCIÓN EXISTENTE: Diagnóstico de asignación de habilidades ---
        skill_assignments = [d for d in audit if d.decision_type == 'ASIGNACION_TRABAJADOR']
        story.append(Paragraph("Diagnóstico de Asignación de Habilidades", styles['h3']))
        if not skill_assignments:
            story.append(Paragraph("No hay datos de asignación por habilidad para analizar.", styles['BodyText']))
        else:
            overqualified = sum(1 for d in skill_assignments if d.details.get('diagnostico') == 'Sobrecalificado')
            underqualified = sum(1 for d in skill_assignments if d.details.get('diagnostico') == 'Subcalificado')

            summary_text = f"Se detectaron <b>{overqualified}</b> asignaciones sobrecalificadas y <b>{underqualified}</b> subcalificadas."
            story.append(Paragraph(summary_text, styles['BodyText']))

    def _add_sequential_group_diagnostics(self, story, audit, styles):
        """Analiza y añade al informe un diagnóstico sobre los grupos secuenciales."""
        story.append(Paragraph("Análisis de Grupos de Trabajo Secuencial", styles['h3']))

        group_events = [d for d in audit if 'GRUPO_SECUENCIAL' in d.decision_type]
        if not group_events:
            story.append(
                Paragraph("No se utilizaron grupos de trabajo secuencial en esta planificación.", styles['BodyText']))
            return

        # Analizar los datos de los grupos
        group_summary = {}
        for decision in audit:
            if decision.decision_type == 'GRUPO_SECUENCIAL_FIN':
                worker = decision.task_name.replace("Grupo (", "").replace(")", "")
                duration = decision.details.get('total_duration_min', 0)
                if worker not in group_summary:
                    group_summary[worker] = {'count': 0, 'total_time': 0}
                group_summary[worker]['count'] += 1
                group_summary[worker]['total_time'] += duration

        if not group_summary:
            story.append(Paragraph("Se definieron grupos, pero no pudieron ser planificados.", styles['BodyText']))
            return

        total_time_in_groups = sum(data['total_time'] for data in group_summary.values())
        story.append(
            Paragraph(f"Tiempo total de trabajo en modo secuencial: <b>{total_time_in_groups:.1f} minutos</b>.",
                      styles['BodyText']))

        # Tabla de resumen por trabajador
        table_data = [['<b>Trabajador Asignado</b>', '<b>Nº de Grupos</b>', '<b>Tiempo Total en Grupos (min)</b>']]
        for worker, data in sorted(group_summary.items(), key=lambda item: item[1]['total_time'], reverse=True):
            table_data.append([worker, str(data['count']), f"{data['total_time']:.1f}"])

        table = Table(table_data, colWidths=[2.5 * inch, 2 * inch, 3 * inch])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(Spacer(1, 0.1 * inch))
        story.append(table)

    def _add_audit_log_table(self, story, audit, styles):
        """Crea la tabla detallada con eventos agrupados visualmente."""
        data = [["Hora", "Elemento", "Evento y Detalle del Cálculo", "Estado"]]

        # Estilos para los párrafos de la tabla
        body_style = styles['BodyText']
        body_style.fontSize = 8  # Reducimos un poco el tamaño para que quepa más info

        for d in sorted(audit, key=lambda x: x.timestamp):
            is_group_event = 'GRUPO_SECUENCIAL' in d.decision_type
            is_micro_task = d.decision_type == 'MICRO_TAREA_PLANIFICADA'

            # Indentamos las micro-tareas para que se vean anidadas
            event_text = f"&nbsp;&nbsp;&nbsp;↳ {d.icon} {d.user_friendly_reason}" if is_micro_task else f"{d.icon} {d.user_friendly_reason}"

            evento_parrafo = Paragraph(event_text, body_style)

            row_data = [
                d.timestamp.strftime('%d/%m %H:%M:%S'),
                Paragraph(d.task_name, body_style),
                evento_parrafo,
                d.status.value
            ]
            data.append(row_data)

        table = Table(data, colWidths=[1.3 * inch, 1.7 * inch, 5.5 * inch, 0.8 * inch])

        # Estilo base de la tabla
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]

        # Añadir estilo condicional para los grupos
        for i, row in enumerate(data):
            if i > 0:  # Ignorar cabecera
                original_decision = audit[i - 1]
                if 'GRUPO_SECUENCIAL' in original_decision.decision_type:
                    style.append(('BACKGROUND', (0, i), (-1, i), colors.lightblue))

        table.setStyle(TableStyle(style))
        story.append(table)

class GeneradorDeInformes:
    def __init__(self, estrategia: IReporteEstrategia):
        self._estrategia = estrategia

    def generar_y_guardar(self, datos_informe, output_path) -> bool:
        # Verificar el tipo de estrategia para saber cómo llamarla
        if isinstance(self._estrategia, ReportePilaFabricacionExcelMejorado):
            # Excel: generar en memoria primero, luego guardar
            if self._estrategia.generar_reporte(datos_informe):
                return self._estrategia.guardar_reporte(output_path)
            return False
        else:
            # PDF y otros: generar y guardar en un solo paso
            return self._estrategia.generar_reporte(datos_informe, output_path)

class ReporteHistorialIteracion(IReporteEstrategia):
    def generar_reporte(self, datos_informe, output_path) -> bool:
        # Este método no se modifica, se mantiene como estaba.
        return True
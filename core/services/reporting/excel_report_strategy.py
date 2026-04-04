"""
Lógica o utilidades del núcleo (`excel_report_strategy`): tipos, servicios auxiliares o infraestructura compartida fuera de la capa de interfaz.
"""
from __future__ import annotations

import logging
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Any, Optional, TYPE_CHECKING
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.services.time_calculator import CalculadorDeTiempos
from core.services.calculation_audit import CalculationDecision
from core.services.report_sheets import (
    ResumenEjecutivoSheet, AnalisisTrabajadoresSheet, CronogramaSheet, 
    CuellosBotollaSheet, AuditSheet, GraficasSheet, TrabajoParaleloSheet
)
from .base import IReporteEstrategia

if TYPE_CHECKING:
    from core.simulation.engine.schedule_config import ScheduleConfig

class ReportePilaFabricacionExcelMejorado(IReporteEstrategia):
    """
    Generador mejorado de reportes Excel con lectura correcta del audit_log
    y presentación clara de grupos secuenciales.
    """

    def __init__(self, schedule_config: Optional[ScheduleConfig] = None) -> None:
        super().__init__()
        self.logger = logging.getLogger(__name__)
        self.schedule_config = schedule_config
        self.time_calculator = CalculadorDeTiempos(self.schedule_config) if schedule_config else None
        self.workbook: Optional[Workbook] = None

    def generar_reporte(self, datos_informe: Dict[Any, Any], output_path: str = "") -> bool:
        """
        Orquesta la creación de todas las hojas del informe en memoria.
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Eliminar la hoja por defecto
            # Guardar datos_informe en el workbook para acceso desde otras funciones
            wb._datos_informe = datos_informe

            all_results = datos_informe.get("data", [])
            audit_log = datos_informe.get("audit_log", [])
            # production_flow = datos_informe.get("production_flow", []) # Removed as it was unused in original

            if not all_results:
                self.logger.warning("No hay datos de simulación para generar el reporte Excel.")
                return False

            self.logger.info(f"Iniciando generación de informe con {len(all_results)} registros cronológicos.")

            # --- INICIO DIAGNÓSTICO CUELLOS DE BOTELLA ---
            self.logger.info(f"DEBUG: Contenido de audit_log recibido ({len(audit_log)} eventos):")
            if not audit_log:
                self.logger.warning("DEBUG: ¡El audit_log está vacío!")
            else:
                for i, event in enumerate(audit_log[:5]):
                    if isinstance(event, CalculationDecision):
                        self.logger.info(
                            f"  Evento {i}: Tipo='{event.decision_type}', Estado='{event.status.value if hasattr(event.status, 'value') else event.status}', Razón='{event.user_friendly_reason[:50]}...'")
                    else:
                        self.logger.info(f"  Evento {i}: (Tipo no reconocido: {type(event)})")
            # --- FIN DIAGNÓSTICO ---

            # 1. Analizar los datos completos una sola vez
            analysis = self._analyze_simulation_data(all_results, audit_log)

            # 2. Crear cada hoja del reporte llamando a los métodos de maquetación
            ResumenEjecutivoSheet().create_sheet(wb, analysis=analysis, datos_informe=datos_informe)
            AnalisisTrabajadoresSheet().create_sheet(wb, all_results=all_results)
            GraficasSheet().create_sheet(wb, all_results=all_results, analysis=analysis)
            CronogramaSheet().create_sheet(wb, all_results=all_results)
            CuellosBotollaSheet().create_sheet(wb, audit_log=audit_log, all_results=all_results)
            TrabajoParaleloSheet().create_sheet(wb, all_results=all_results)
            AuditSheet().create_sheet(wb, audit_log=audit_log)

            self.workbook = wb
            self.logger.info("Informe Excel generado en memoria con todas las hojas y maquetación.")
            return True

        except Exception as e:
            self.logger.error(f"Error crítico durante la generación del reporte Excel: {e}", exc_info=True)
            return False

    def _crear_hoja_grupos_secuenciales(self, wb: Workbook, all_results: List[Dict[str, Any]], production_flow: List[Dict[str, Any]]) -> None:
        """Crea una hoja con el detalle de los grupos secuenciales planificados."""
        ws = wb.create_sheet("Grupos Secuenciales")
        ws['A1'] = "DETALLE DE GRUPOS SECUENCIALES"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:G1')

        row = 3
        headers = ["ID Grupo", "Trabajador", "Tareas", "Inicio", "Fin", "Duración (min)", "Estado"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row += 1

        if not production_flow:
            ws[f'A{row}'] = "No hay información de grupos secuenciales disponible"
            ws[f'A{row}'].font = Font(italic=True)
            return

        # Logic to extract group info from production_flow and results
        # (Simplified implementation for now to satisfy tests)
        for i, section in enumerate(production_flow):
            if section.get('type') == 'sequential_group':
                worker = section.get('assigned_worker', 'N/A')
                tasks = [t.get('task', {}).get('name', 'N/A') for t in section.get('tasks', [])]
                
                # Find timing in results
                group_tasks_results = [r for r in all_results if r.get('Tarea') in tasks]
                if group_tasks_results:
                    inicio = min(r['Inicio'] for r in group_tasks_results)
                    fin = max(r['Fin'] for r in group_tasks_results)
                    duracion = sum(r['Duracion (min)'] for r in group_tasks_results)
                else:
                    inicio = fin = duracion = "N/A"

                ws[f'A{row}'] = f"GRP_{i+1:03d}"
                ws[f'B{row}'] = worker
                ws[f'C{row}'] = ", ".join(tasks)
                ws[f'D{row}'] = inicio.strftime('%d/%m/%Y %H:%M') if isinstance(inicio, datetime) else str(inicio)
                ws[f'E{row}'] = fin.strftime('%d/%m/%Y %H:%M') if isinstance(fin, datetime) else str(fin)
                ws[f'F{row}'] = duracion
                ws[f'G{row}'] = "Planificado"
                row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

    def _analyze_simulation_data(self, results: List[Dict[str, Any]], audit_log: List[Any]) -> Dict[str, Any]:
        analysis = {
            'total_tasks': len(results),
            'total_duration_min': sum(r['Duracion (min)'] for r in results),
            'start_time': min(r['Inicio'] for r in results) if results else None,
            'end_time': max(r['Fin'] for r in results) if results else None,
            'workers_involved': set(),
            'machines_used': set(),
            'departments': defaultdict(float),
            'idle_times': [],
            'bottlenecks': [],
            'groups_performance': []
        }
        for result in results:
            workers = result.get('Trabajador Asignado', [])
            if isinstance(workers, list):
                valid_workers = [w for w in workers if w and str(w).strip()]
                analysis['workers_involved'].update(valid_workers)
            elif isinstance(workers, str):
                if ',' in workers:
                    worker_list = [w.strip() for w in workers.split(',') if w.strip()]
                    analysis['workers_involved'].update(worker_list)
                elif workers and workers.strip():
                    analysis['workers_involved'].add(workers.strip())
            machine = result.get('nombre_maquina')
            if machine and machine != 'N/A':
                analysis['machines_used'].add(machine)
            dept = result.get('Departamento', 'General')
            analysis['departments'][dept] += result['Duracion (min)']

        for decision in audit_log:
            if hasattr(decision, 'decision_type'):
                if decision.decision_type == 'TIEMPO_DE_ESPERA':
                    analysis['idle_times'].append({
                        'task': decision.task_name,
                        'duration': decision.details.get('wait_time', 0),
                        'reason': decision.reason
                    })
                elif decision.decision_type == 'CUELLO_DE_BOTELLA':
                    analysis['bottlenecks'].append({
                        'resource': decision.details.get('resource'),
                        'impact': decision.details.get('impact_minutes', 0),
                        'affected_tasks': decision.details.get('affected_tasks', [])
                    })

                instancias_encontradas = set()
                tareas_con_paralelo = defaultdict(set)
                for task in results:
                    inst_id = task.get('Instancia ID')
                    if inst_id and inst_id != 'N/A' and inst_id != 'Principal':
                        instancias_encontradas.add(inst_id)
                        tarea_nombre = task.get('Tarea', 'Desconocida')
                        tareas_con_paralelo[tarea_nombre].add(inst_id)
                total_instancias_paralelas = 0
                max_instancias_simultaneas_en_tarea = 0
                for tarea, instancias_set in tareas_con_paralelo.items():
                    num_instancias = len(instancias_set)
                    if num_instancias > 1:
                        total_instancias_paralelas += num_instancias
                        if num_instancias > max_instancias_simultaneas_en_tarea:
                            max_instancias_simultaneas_en_tarea = num_instancias
                if total_instancias_paralelas == 0 and len(instancias_encontradas) > 0:
                    total_instancias_paralelas = len(instancias_encontradas)
                if max_instancias_simultaneas_en_tarea == 0 and len(instancias_encontradas) > 0:
                    max_instancias_simultaneas_en_tarea = max(len(s) for s in tareas_con_paralelo.values()) if tareas_con_paralelo else 0
                analysis['total_instancias_paralelas'] = total_instancias_paralelas
                analysis['max_instancias_simultaneas'] = max_instancias_simultaneas_en_tarea
        return analysis

    def guardar_reporte(self, output_path: str) -> bool:
        if not self.workbook:
            self.logger.error("No hay un workbook para guardar. Ejecute generar_reporte() primero.")
            return False
        try:
            self.workbook.save(output_path)
            self.logger.info(f"Reporte Excel guardado en: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Error al guardar el archivo Excel: {e}", exc_info=True)
            return False

    def _crear_hoja_trabajo_paralelo(self, wb: Workbook, all_results: list[dict[str, Any]]) -> None:
        ws = wb.create_sheet("Trabajo Paralelo")
        ws['A1'] = "ANÁLISIS DE TRABAJO PARALELO POR INSTANCIA"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:H1')
        resultados_paralelos = []
        for res in all_results:
            inst_id = res.get('Instancia ID')
            if inst_id and inst_id != 'N/A' and inst_id != 'Principal':
                resultados_paralelos.append(res)
        if not resultados_paralelos:
            ws['A3'] = "No se detectó trabajo paralelo (instancias múltiples) en esta simulación."
            ws['A3'].font = Font(italic=True)
            return
        instancias = defaultdict(list)
        for res in resultados_paralelos:
            inst_id = res['Instancia ID']
            instancias[inst_id].append(res)
        headers = ["Tarea", "Instancia ID (8 dígitos)", "Trabajadores en Instancia", "Unidad Inicial", "Unidad Final", "Inicio", "Fin", "Duración Total (min)"]
        current_row = 3
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        for inst_id, registros in sorted(instancias.items(), key=lambda item: (item[1][0].get('Tarea', ''), item[0])):
            registros.sort(key=lambda x: x.get('Inicio', datetime.min))
            tarea = registros[0].get('Tarea', 'N/A')
            trabajadores = registros[0].get('Lista Trabajadores', [])
            unidad_inicial = min(int(r.get('Numero Unidad', 0) or 0) for r in registros)
            unidad_final = max(int(r.get('Numero Unidad', 0) or 0) for r in registros)
            inicios: list[datetime] = [
                v for r in registros for v in (r.get('Inicio'),) if isinstance(v, datetime)
            ]
            fines: list[datetime] = [
                v for r in registros for v in (r.get('Fin'),) if isinstance(v, datetime)
            ]
            inicio = min(inicios) if inicios else None
            fin = max(fines) if fines else None
            duracion_total = sum(r.get('Duracion (min)', 0) for r in registros)
            row_data = [tarea, inst_id[:8], ", ".join(trabajadores), unidad_inicial, unidad_final, inicio.strftime('%d/%m/%Y %H:%M') if inicio else 'N/A', fin.strftime('%d/%m/%Y %H:%M') if fin else 'N/A', round(duracion_total, 2)]
            ws.append(row_data)
            if (current_row - 3) % 2 == 0:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        column_widths = [30, 18, 30, 12, 12, 20, 20, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A4"
        self.logger.info(f"Hoja 'Trabajo Paralelo' creada con {len(instancias)} instancias detalladas.")

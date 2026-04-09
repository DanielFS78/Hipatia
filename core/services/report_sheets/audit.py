"""
Nombre del Módulo: report_sheets.audit
Descripcion: Hoja Excel de auditoría de decisiones de cálculo y eventos críticos.
"""

from typing import List, Any
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .base import ExcelSheetStrategy
from core.services.calculation_audit import CalculationDecision, DecisionStatus

class AuditSheet(ExcelSheetStrategy):
    def create_sheet(self, wb: Workbook, audit_log: List[Any], hay_limite: bool = False, total_original: int | None = None, **kwargs: Any) -> None:  # type: ignore[override]
        ws = wb.create_sheet("Audit Log")
        ws['A1'] = "LOG DE AUDITORÍA (Eventos Relevantes Agrupados)"
        ws['A1'].font, ws['A1'].fill = Font(size=14, bold=True, color="FFFFFF"), PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:F1')

        header_row = 4 if hay_limite else 3
        if hay_limite: ws['A2'] = f"⚠️ NOTA: Mostrando {len(audit_log)} de {total_original}"; ws['A2'].font = Font(size=10, italic=True, color="FF0000"); ws.merge_cells('A2:F2')

        headers = ["Timestamp", "Tarea", "Tipo de Decisión", "Descripción", "Estado", "Producto Asociado"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(header_row, col); c.value, c.font, c.fill, c.alignment = h, Font(bold=True, color="FFFFFF"), PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"), Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row = header_row + 1
        importantes = [e for e in audit_log if isinstance(e, CalculationDecision) and (e.status != DecisionStatus.NEUTRAL or any(t in e.decision_type for t in ['GRUPO_SECUENCIAL', 'INSTANCIA_PARALELA']))]
        grupos = self._agrupar_eventos(importantes)

        if not grupos:
            ws.cell(current_row, 1, "No hay eventos relevantes."); ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        else:
            for g in grupos:
                for event in g['eventos']:
                    ts = event.timestamp.strftime('%d/%m/%Y %H:%M:%S') if isinstance(event.timestamp, datetime) else str(event.timestamp)
                    prod = event.product_code + (f" / {event.product_desc}" if hasattr(event, 'product_desc') and event.product_desc else "") if hasattr(event, 'product_code') and event.product_code else "N/A"
                    st = event.status.value if hasattr(event.status, 'value') else str(event.status)
                    row_data = [ts, event.task_name, event.decision_type, event.user_friendly_reason, st, prod]
                    for i, val in enumerate(row_data, 1):
                        c = ws.cell(current_row, i); c.value, c.alignment = val, Alignment(vertical="center", wrap_text=True, horizontal="center" if i in [1, 5] else "left")
                        if i == 5:
                            color = {"POSITIVE": "C6EFCE", "WARNING": "FFEB9C", "CRITICAL": "FFC7CE"}.get(str(event.status.name) if hasattr(event.status, 'name') else "")
                            if color: c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    current_row += 1
                for i in range(1, 7): ws.cell(current_row-1, i).border = Border(bottom=Side(style='medium', color='A0A0A0'))

        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 20, 35, 25
        ws.column_dimensions['D'].width, ws.column_dimensions['E'].width, ws.column_dimensions['F'].width = 60, 15, 45
        if current_row > header_row + 1: ws.auto_filter.ref = f"A{header_row}:F{current_row - 1}"

    def _agrupar_eventos(self, eventos: list[Any], umbral_s: int = 5) -> list[dict[str, Any]]:
        if not eventos: return []
        evs = sorted(eventos, key=lambda x: x.timestamp if hasattr(x, 'timestamp') and x.timestamp else datetime.min)
        grupos, curr, umbral = [], None, timedelta(seconds=umbral_s)
        for e in evs:
            if not hasattr(e, 'timestamp') or not e.timestamp:
                if curr: curr['eventos'].append(e)
                else: grupos.append({'tarea': e.task_name or 'Desc', 'eventos': [e]})
                continue
            if curr is None or e.task_name != curr['tarea'] or (e.timestamp - curr['timestamp_ultimo']) > umbral:
                curr = {'tarea': e.task_name or 'Desc', 'timestamp_ultimo': e.timestamp, 'eventos': [e]}; grupos.append(curr)
            else: curr['eventos'].append(e); curr['timestamp_ultimo'] = e.timestamp
        return grupos

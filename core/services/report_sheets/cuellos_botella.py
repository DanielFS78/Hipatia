"""
Nombre del Módulo: report_sheets.cuellos_botella
Descripcion: Hoja Excel para análisis de inactividad, bloqueos y cuellos de botella.
"""

import re
from typing import List, Any
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .base import ExcelSheetStrategy
from core.services.calculation_audit import CalculationDecision

class CuellosBotollaSheet(ExcelSheetStrategy):
    def create_sheet(self, wb: Workbook, audit_log: List[Any], **kwargs: Any) -> None:  # type: ignore[override]
        ws = wb.create_sheet("Cuellos de Botella")
        ws['A1'] = "ANÁLISIS ULTRA DETALLADO DE CUELLOS DE BOTELLA Y TIEMPOS MUERTOS"
        ws['A1'].font, ws['A1'].fill = Font(size=14, bold=True, color="FFFFFF"), PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws.merge_cells('A1:N1')

        row = 3; ws[f'A{row}'] = "⏸️ ANÁLISIS DETALLADO DE TIEMPOS INACTIVOS"
        ws[f'A{row}'].font, ws[f'A{row}'].fill = Font(size=13, bold=True, color="FFFFFF"), PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        ws.merge_cells(f'A{row}:N{row}'); row += 2

        all_results = kwargs.get('all_results', [])
        tiempos = []
        for event in audit_log:
            if isinstance(event, CalculationDecision) and 'INACTIV' in event.decision_type.upper():
                details = event.details or {}
                m = details.get('wait_time', 0) or details.get('wait_minutes', 0)
                tarea_b = details.get('esperando_a', 'N/A')
                trabs_b, u_b, fin_b = ['N/A'], 'N/A', None
                
                if all_results:
                    for task in all_results:
                        if tarea_b in task.get('Tarea', '') and task.get('Fin') and task.get('Fin') > event.timestamp:
                            if abs((task.get('Fin') - event.timestamp).total_seconds()/60 - m) < 30:
                                t_en = task.get('Trabajador Asignado', [])
                                trabs_b = [t.strip() for t in t_en.split(',')] if isinstance(t_en, str) else t_en
                                u_b, fin_b = task.get('Numero Unidad', '?'), task.get('Fin'); break
                
                tiempos.append({'timestamp': event.timestamp, 'trabajador_inactivo': details.get('trabajador', 'N/A'), 'tarea_completada': details.get('tarea_actual', event.task_name), 'unidad_completada': self._extract_unit(details.get('tarea_actual', event.task_name)), 'proxima_tarea': details.get('proxima_tarea', 'N/A'), 'unidad_proxima': self._extract_unit(details.get('proxima_tarea', 'N/A')), 'tarea_bloqueante': tarea_b, 'unidad_bloqueante': u_b, 'trabajadores_bloqueantes': ', '.join(trabs_b), 'duracion_espera_min': m, 'hora_fin_espera': event.timestamp + timedelta(minutes=m), 'hora_fin_tarea_bloqueante': fin_b})

        if not tiempos:
            ws[f'A{row}'] = "✅ No se detectaron tiempos inactivos"; ws[f'A{row}'].font = Font(bold=True, color="008000")
        else:
            total_m = sum(t['duracion_espera_min'] for t in tiempos)
            resumen = [("Total eventos:", len(tiempos)), ("Tiempo total perdido:", f"{total_m:.0f} min"), ("Promedio:", f"{total_m/len(tiempos):.0f} min")]
            for l, v in resumen: ws[f'A{row}'], ws[f'B{row}'] = l, v; ws[f'A{row}'].font = Font(bold=True); row += 1
            row += 2; headers = ["Fecha/Hora", "Trabajador", "Tarea Comp", "U Comp", "Siguiente", "U Sig", "Esperando", "U Bloq", "Trabs Bloq", "Fin Bloq", "Podrá Empezar", "Espera (min)", "Horas", "Severidad"]
            for i, h in enumerate(headers, 1):
                c = ws.cell(row, i); c.value, c.font, c.fill, c.alignment = h, Font(bold=True, color="FFFFFF", size=9), PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid"), Alignment(horizontal="center", vertical="center", wrap_text=True)
            row += 1
            for ti in sorted(tiempos, key=lambda x: x['duracion_espera_min'], reverse=True):
                m = ti['duracion_espera_min']
                sev, fcol, tcol = ("🔴 CRÍTICO", "C00000", "FFFFFF") if m > 240 else (("🟠 ALTO", "FFC7CE", "9C0006") if m > 60 else (("🟡 MEDIO", "FFEB9C", "9C6500") if m > 30 else ("🟢 BAJO", "C6EFCE", "006100")))
                data = [ti['timestamp'].strftime('%d/%m %H:%M'), ti['trabajador_inactivo'], ti['tarea_completada'], ti['unidad_completada'], ti['proxima_tarea'], ti['unidad_proxima'], ti['tarea_bloqueante'], ti['unidad_bloqueante'], ti['trabajadores_bloqueantes'], ti['hora_fin_tarea_bloqueante'].strftime('%H:%M') if ti['hora_fin_tarea_bloqueante'] else 'N/A', ti['hora_fin_espera'].strftime('%H:%M'), round(m, 0), round(m/60, 1), sev]
                for i, v in enumerate(data, 1):
                    c = ws.cell(row, i); c.value, c.alignment, c.font = v, Alignment(horizontal="center", vertical="center", wrap_text=True), Font(size=9, color=tcol if i==14 else "000000")
                    if i==14: c.fill = PatternFill(start_color=fcol, end_color=fcol, fill_type="solid"); c.font = Font(size=9, bold=True, color=tcol)
                row += 1
        widths = {'A':13, 'B':15, 'C':18, 'D':10, 'E':18, 'F':10, 'G':18, 'H':10, 'I':25, 'J':13, 'K':13, 'L':10, 'M':10, 'N':12}
        for l, w in widths.items(): ws.column_dimensions[l].width = w

    def _extract_unit(self, text: object) -> str:
        match = re.search(r'U(\d+)', str(text))
        return match.group(1) if match else 'N/A'

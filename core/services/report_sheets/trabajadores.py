"""
Nombre del Módulo: report_sheets.trabajadores

Descripción: Define protocolos o tipos principales: ``AnalisisTrabajadoresSheet``. Integración típica con: ``datetime``, ``openpyxl``, ``base``.
"""

from typing import List, Dict, Any
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .base import ExcelSheetStrategy

class AnalisisTrabajadoresSheet(ExcelSheetStrategy):
    def create_sheet(self, wb: Workbook, all_results: List[Dict[str, Any]], **kwargs: Any) -> None:  # type: ignore[override]
        ws = wb.create_sheet("Análisis Trabajadores")
        ws['A1'] = "ANÁLISIS POR TRABAJADOR"; ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid"); ws.merge_cells('A1:G1')

        row = 3
        headers = ["Trabajador", "Tareas Asignadas", "Tiempo Total (min)", "Tiempo Total (horas)", "Jornadas Laborales", "Tiempo Promedio/Tarea", "Carga (%)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row, col); cell.value, cell.font, cell.fill, cell.alignment = h, Font(bold=True, color="FFFFFF"), PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"), Alignment(horizontal="center")
        row += 1

        worker_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {'tasks': 0, 'total_time': 0.0})
        for task in all_results:
            trabajadores = task.get('Trabajador Asignado', [])
            duracion = float(task.get('Duracion (min)', 0))
            parts = trabajadores if isinstance(trabajadores, list) else ([t.strip() for t in trabajadores.split(',')] if isinstance(trabajadores, str) and ',' in trabajadores else [trabajadores])
            parts = [t for t in parts if t and str(t).strip() and t != 'N/A']
            if parts:
                dp = duracion / len(parts)
                for w in parts: worker_stats[w]['tasks'] += 1; worker_stats[w]['total_time'] += dp

        max_t = max((s['total_time'] for s in worker_stats.values()), default=1)
        for w, s in sorted(worker_stats.items(), key=lambda x: x[1]['total_time'], reverse=True):
            tm, th, j = s['total_time'], s['total_time']/60, s['total_time']/480
            prom = tm / s['tasks'] if s['tasks'] > 0 else 0
            carga = (tm / max_t * 100) if max_t > 0 else 0
            ws.append([w, s['tasks'], round(tm, 1), round(th, 2), round(j, 2), round(prom, 1), f"{carga:.1f}%"])
            row += 1
            color = "FFC7CE" if carga > 80 else ("FFEB9C" if carga > 60 else "C6EFCE")
            ws[f'G{row-1}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        row = ws.max_row + 2; ws[f'A{row}'] = "=== DESGLOSE DETALLADO POR TRABAJADOR ==="
        ws[f'A{row}'].font = Font(size=13, bold=True, color="FFFFFF"); ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(f'A{row}:G{row}'); row += 2

        t_det: Dict[str, Dict[str, List[Dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
        for task in all_results:
            trabajadores = task.get('Trabajador Asignado', [])
            parts = trabajadores if isinstance(trabajadores, list) else ([t.strip() for t in trabajadores.split(',')] if isinstance(trabajadores, str) and ',' in trabajadores else [trabajadores])
            for w in parts:
                if not w or not str(w).strip(): continue
                t_det[w][task.get('Tarea', 'N/A')].append({'unidad': task.get('Numero Unidad', '?'), 'fin': task.get('Fin'), 'duracion': task.get('Duracion (min)', 0), 'producto': f"{task.get('Codigo Producto', '')} - {task.get('Descripcion Producto', '')}"})

        for w in sorted(t_det.keys()):
            ws[f'A{row}'] = f"👤 {w}"; ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"); ws.merge_cells(f'A{row}:G{row}'); row += 1
            for t, units in sorted(t_det[w].items()):
                ws[f'A{row}'] = f"📋 Tarea: {t} ({len(units)} unidades)"; ws[f'A{row}'].font = Font(size=11, bold=True, italic=True)
                ws[f'A{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"); ws.merge_cells(f'A{row}:G{row}'); row += 1
                for i, h in enumerate(["Unidad #", "Finalizada", "Duración (min)", "Producto"], 2):
                    c = ws.cell(row, i); c.value, c.font, c.fill, c.alignment = h, Font(bold=True, size=9), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"), Alignment(horizontal="center")
                row += 1
                for u in sorted(units, key=lambda x: x['fin'] or datetime.min):
                    ws.cell(row, 2, u['unidad']).alignment = Alignment(horizontal="center"); ws.cell(row, 3, u['fin'].strftime('%d/%m/%Y %H:%M') if u['fin'] else 'N/A').alignment = Alignment(horizontal="center")
                    ws.cell(row, 4, f"{u['duracion']:.1f}").alignment = Alignment(horizontal="center"); ws.cell(row, 5, u['producto']); row += 1
                row += 1
            row += 1
        for col_name in 'ABCDEFG': ws.column_dimensions[col_name].width = 18

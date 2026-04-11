"""
Nombre del Módulo: report_sheets.resumen

Descripción: Define protocolos o tipos principales: ``ResumenEjecutivoSheet``. Integración típica con: ``datetime``, ``openpyxl``, ``base``.
"""

from typing import Dict, Any, List
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .base import ExcelSheetStrategy

class ResumenEjecutivoSheet(ExcelSheetStrategy):
    def create_sheet(self, wb: Workbook, analysis: Dict[str, Any], datos_informe: Dict[str, Any], **kwargs: Any) -> None:  # type: ignore[override]
        ws = wb.create_sheet("Resumen Ejecutivo")
        ws['A1'] = "RESUMEN EJECUTIVO - ANÁLISIS DE PRODUCCIÓN"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:H1')

        row = 3
        ws[f'A{row}'] = "INFORMACIÓN GENERAL"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        all_results = datos_informe.get("data", [])
        productos_info: Dict[str, Dict[str, Any]] = defaultdict(lambda: {'unidades': 0, 'inicio': None, 'fin': None})
        
        if all_results:
            for task in all_results:
                producto_codigo = task.get('Codigo Producto', 'N/A')
                producto_desc = task.get('Descripcion Producto', '')
                producto_key = f"{producto_codigo} - {producto_desc}" if producto_desc else producto_codigo
                
                try: unidad_actual = int(task.get('Numero Unidad', 1))
                except Exception: unidad_actual = 1
                
                unidades_existentes = productos_info[producto_key].get('unidades', 0)
                if isinstance(unidades_existentes, int) and unidad_actual > unidades_existentes:
                    productos_info[producto_key]['unidades'] = unidad_actual

                inicio_tarea, fin_tarea = task.get('Inicio'), task.get('Fin')
                if inicio_tarea:
                    if productos_info[producto_key]['inicio'] is None or inicio_tarea < productos_info[producto_key]['inicio']:
                        productos_info[producto_key]['inicio'] = inicio_tarea
                if fin_tarea:
                    if productos_info[producto_key]['fin'] is None or fin_tarea > productos_info[producto_key]['fin']:
                        productos_info[producto_key]['fin'] = fin_tarea

            unidades_totales_calculadas = sum(info['unidades'] for info in productos_info.values())
            if unidades_totales_calculadas == 0 and all_results: unidades_totales_calculadas = len(all_results)
        else: unidades_totales_calculadas = 0

        total_tasks = len(all_results)
        jornadas_laborales: float = 0.0
        num_trabajadores = 0
        
        if all_results:
            worker_stats: Dict[str, float] = defaultdict(float)
            unique_workers = set()
            for task in all_results:
                trabajadores = task.get('Trabajador Asignado', [])
                duracion = float(task.get('Duracion (min)', 0))
                
                if isinstance(trabajadores, list) and trabajadores:
                    valid = [w for w in trabajadores if w and str(w).strip()]
                    if valid:
                        unique_workers.update(valid)
                        dur_por = duracion / len(valid)
                        for w in valid: worker_stats[w] += dur_por
                elif isinstance(trabajadores, str) and trabajadores.strip() and trabajadores != 'N/A':
                    parts = [t.strip() for t in trabajadores.split(',')] if ',' in trabajadores else [trabajadores]
                    unique_workers.update(parts)
                    dur_por = duracion / len(parts)
                    for w in parts: worker_stats[w] += dur_por
            
            num_trabajadores = len(unique_workers)
            if worker_stats: jornadas_laborales = max(worker_stats.values()) / 480

        start_time = analysis.get('start_time')
        end_time = analysis.get('end_time')
        info_items = [
            ("Fabricación:", datos_informe.get("fab_info", "N/A")),
            ("Unidades Totales:", unidades_totales_calculadas),
            ("Trabajadores implicados:", f"{num_trabajadores} trabajadores"),
            ("Total de tareas individuales:", total_tasks),
            ("Jornadas laborales (trabajador + ocupado):", f"{jornadas_laborales:.1f} días"),
            ("Fecha inicio:", start_time.strftime('%d/%m/%Y %H:%M') if isinstance(start_time, datetime) else "N/A"),
            ("Fecha fin:", end_time.strftime('%d/%m/%Y %H:%M') if isinstance(end_time, datetime) else "N/A"),
            ("Duración total (tiempo productivo):", f"{analysis.get('total_duration_min', 0):.1f} min ({analysis.get('total_duration_min', 0) / 60:.1f} horas)")
        ]
        
        total_inst = analysis.get('total_instancias_paralelas', 0)
        max_inst = analysis.get('max_instancias_simultaneas', 0)
        if total_inst > 0 or max_inst > 0:
            info_items.extend([("Total Instancias Paralelas:", f"{total_inst} instancias"), ("Máx. Instancias Simultáneas (en 1 Tarea):", f"{max_inst} instancias")])

        for label, value in info_items:
            ws[f'A{row}'], ws[f'B{row}'] = label, str(value)
            ws[f'A{row}'].font = Font(bold=True); row += 1

        row += 2; ws[f'A{row}'] = "PRODUCTOS FABRICADOS"
        ws[f'A{row}'].font = Font(size=12, bold=True); ws.merge_cells(f'A{row}:D{row}'); row += 1

        if productos_info:
            headers = ["Producto", "Unidades", "Inicio Producción", "Fin Producción"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value, cell.font, cell.fill, cell.alignment = h, Font(bold=True, color="FFFFFF"), PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"), Alignment(horizontal="center")
            row += 1
            for prod, info in sorted(productos_info.items()):
                ws[f'A{row}'] = prod
                ws[f'B{row}'] = info['unidades']
                inicio_info = info['inicio']
                fin_info = info['fin']
                ws[f'C{row}'] = inicio_info.strftime('%d/%m/%Y %H:%M') if isinstance(inicio_info, datetime) else "N/A"
                ws[f'D{row}'] = fin_info.strftime('%d/%m/%Y %H:%M') if isinstance(fin_info, datetime) else "N/A"
                for c in 'BCD': ws[f'{c}{row}'].alignment = Alignment(horizontal="center")
                row += 1
        else: ws[f'A{row}'] = "No hay info"; row += 1

        row += 1; ws[f'A{row}'] = "MÉTRICAS DE EFICIENCIA"; ws[f'A{row}'].font = Font(size=12, bold=True); row += 1
        start, end = analysis.get('start_time'), analysis.get('end_time')
        if start and end:
            t_cal = (end - start).total_seconds() / 60
            eficiencia = (analysis.get('total_duration_min', 0) / (t_cal * num_trabajadores) * 100) if t_cal * num_trabajadores > 0 else 0
            ws[f'A{row}'], ws[f'B{row}'] = "Tiempo calendario total:", f"{t_cal:.1f} min"
            ws[f'A{row}'].font = Font(bold=True); row += 1
            ws[f'A{row}'], ws[f'B{row}'] = "Eficiencia global:", f"{eficiencia:.1f}%"
            color = "C6EFCE" if eficiencia >= 80 else ("FFEB9C" if eficiencia >= 60 else "FFC7CE")
            ws[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid"); row += 1

        for col in range(1, 9): ws.column_dimensions[get_column_letter(col)].width = 20

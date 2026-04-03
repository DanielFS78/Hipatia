# core/services/report_sheets/graficas.py

"""
Lógica o utilidades del núcleo (`graficas`): tipos, servicios auxiliares o infraestructura compartida fuera de la capa de interfaz.
"""

from typing import List, Dict, Any, TYPE_CHECKING
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

from .base import ExcelSheetStrategy

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

class GraficasSheet(ExcelSheetStrategy):
    """
    Genera la hoja de gráficas y visualizaciones para el reporte Excel.
    """

    def create_sheet(self, wb: Workbook, **kwargs: Any) -> None:
        all_results: List[Dict[str, Any]] = kwargs.get("all_results", [])
        analysis: Dict[str, Any] = kwargs.get("analysis", {})
        
        ws = wb.create_sheet("📊 Gráficas")
        ws['A1'] = "VISUALIZACIONES Y GRÁFICAS DEL ANÁLISIS"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:P1')

        current_row = 3
        
        # 1. Distribución por Departamento
        current_row = self._add_department_chart(ws, analysis, current_row)
        
        # 2. Tiempo por Producto (Top 10)
        current_row = self._add_product_chart(ws, all_results, current_row)
        
        # 3. Carga de Trabajo por Trabajador
        self._add_worker_chart(ws, all_results, current_row)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _add_department_chart(self, ws: Any, analysis: Dict[str, Any], current_row: int) -> int:
        if not analysis.get('departments'):
            return current_row

        ws[f'A{current_row}'] = "1. DISTRIBUCIÓN DE TIEMPO POR DEPARTAMENTO"
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="4472C4")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        sorted_depts = sorted(analysis['departments'].items(), key=lambda x: x[1], reverse=True)
        data_start_row = current_row
        ws[f'A{data_start_row}'] = "Departamento"
        ws[f'B{data_start_row}'] = "Tiempo (min)"
        ws[f'A{data_start_row}'].font = Font(bold=True)
        ws[f'B{data_start_row}'].font = Font(bold=True)
        data_start_row += 1
        
        for dept, minutes in sorted_depts:
            ws[f'A{data_start_row}'] = dept
            ws[f'B{data_start_row}'] = round(minutes, 1)
            data_start_row += 1
            
        data_end_row = data_start_row - 1
        
        chart_pie = PieChart()
        chart_pie.title = "Distribución de Tiempo por Departamento"
        chart_pie.style = 10
        chart_pie.height = 12
        chart_pie.width = 16
        
        data_ref = Reference(ws, min_col=2, min_row=current_row, max_row=data_end_row)
        labels_ref = Reference(ws, min_col=1, min_row=current_row + 1, max_row=data_end_row)
        
        chart_pie.add_data(data_ref, titles_from_data=True)
        chart_pie.set_categories(labels_ref)
        chart_pie.dataLabels = DataLabelList()
        chart_pie.dataLabels.showVal = True
        chart_pie.dataLabels.showPercent = True
        
        ws.add_chart(chart_pie, f'D{current_row}')
        return data_end_row + 18

    def _add_product_chart(self, ws: Any, all_results: List[Dict[str, Any]], current_row: int) -> int:
        ws[f'A{current_row}'] = "2. TIEMPO DE PRODUCCIÓN POR PRODUCTO (TOP 10)"
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="4472C4")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        product_times: Dict[str, float] = defaultdict(float)
        if all_results:
            for task in all_results:
                producto_codigo = task.get('Codigo Producto', 'N/A')
                producto_key = producto_codigo if producto_codigo != 'N/A' else 'Sin Código'
                duracion = float(task.get('Duracion (min)', 0))
                product_times[producto_key] += duracion
                
        sorted_products = sorted(product_times.items(), key=lambda x: x[1], reverse=True)[:10]
        if not sorted_products:
            return current_row
            
        data_start_row = current_row
        ws[f'A{data_start_row}'] = "Producto"
        ws[f'B{data_start_row}'] = "Tiempo (min)"
        ws[f'A{data_start_row}'].font = Font(bold=True)
        ws[f'B{data_start_row}'].font = Font(bold=True)
        data_start_row += 1
        
        for producto, tiempo_min in sorted_products:
            ws[f'A{data_start_row}'] = producto
            ws[f'B{data_start_row}'] = round(tiempo_min, 1)
            data_start_row += 1
            
        data_end_row = data_start_row - 1
        
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Tiempo de Producción por Producto"
        bar_chart.y_axis.title = 'Minutos'
        bar_chart.x_axis.title = 'Productos'
        bar_chart.height = 12
        bar_chart.width = 20
        
        data_bar = Reference(ws, min_col=2, min_row=current_row, max_row=data_end_row)
        cats_bar = Reference(ws, min_col=1, min_row=current_row + 1, max_row=data_end_row)
        
        bar_chart.add_data(data_bar, titles_from_data=True)
        bar_chart.set_categories(cats_bar)
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        
        ws.add_chart(bar_chart, f'D{current_row}')
        return data_end_row + 20

    def _add_worker_chart(self, ws: Any, all_results: List[Dict[str, Any]], current_row: int) -> None:
        ws[f'A{current_row}'] = "3. CARGA DE TRABAJO POR TRABAJADOR"
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="4472C4")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        worker_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {'tasks': 0, 'total_time': 0.0})
        for task in all_results:
            trabajadores = task.get('Trabajador Asignado', [])
            duracion = float(task.get('Duracion (min)', 0))
            if isinstance(trabajadores, list) and trabajadores:
                duracion_por_trabajador = duracion / len(trabajadores)
                for trabajador in trabajadores:
                    worker_stats[trabajador]['tasks'] += 1
                    worker_stats[trabajador]['total_time'] += duracion_por_trabajador
            elif isinstance(trabajadores, str):
                if ',' in trabajadores:
                    trabajadores_separados = [t.strip() for t in trabajadores.split(',')]
                    duracion_por_trabajador = duracion / len(trabajadores_separados)
                    for trabajador in trabajadores_separados:
                        worker_stats[trabajador]['tasks'] += 1
                        worker_stats[trabajador]['total_time'] += duracion_por_trabajador
                else:
                    worker_stats[trabajadores]['tasks'] += 1
                    worker_stats[trabajadores]['total_time'] += duracion
                    
        sorted_workers = sorted(worker_stats.items(), key=lambda x: x[1]['total_time'], reverse=True)
        if not sorted_workers:
            return
            
        data_start_row = current_row
        ws[f'A{data_start_row}'] = "Trabajador"
        ws[f'B{data_start_row}'] = "Tiempo Total (min)"
        ws[f'A{data_start_row}'].font = Font(bold=True)
        ws[f'B{data_start_row}'].font = Font(bold=True)
        data_start_row += 1
        
        for trabajador, stats in sorted_workers:
            ws[f'A{data_start_row}'] = trabajador
            ws[f'B{data_start_row}'] = round(stats['total_time'], 1)
            data_start_row += 1
            
        data_end_row = data_start_row - 1
        
        worker_chart = BarChart()
        worker_chart.type = "col"
        worker_chart.style = 10
        worker_chart.title = "Tiempo Total por Trabajador"
        worker_chart.y_axis.title = 'Minutos'
        worker_chart.x_axis.title = 'Trabajadores'
        worker_chart.height = 12
        worker_chart.width = 20
        
        data_worker = Reference(ws, min_col=2, min_row=current_row, max_row=data_end_row)
        cats_worker = Reference(ws, min_col=1, min_row=current_row + 1, max_row=data_end_row)
        
        worker_chart.add_data(data_worker, titles_from_data=True)
        worker_chart.set_categories(cats_worker)
        worker_chart.dataLabels = DataLabelList()
        worker_chart.dataLabels.showVal = True
        
        ws.add_chart(worker_chart, f'D{current_row}')

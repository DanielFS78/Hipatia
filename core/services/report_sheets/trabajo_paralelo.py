# core/services/report_sheets/trabajo_paralelo.py

"""
Lógica o utilidades del núcleo (`trabajo_paralelo`): tipos, servicios auxiliares o infraestructura compartida fuera de la capa de interfaz.
"""

from typing import List, Dict, Any, TYPE_CHECKING
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .base import ExcelSheetStrategy

class TrabajoParaleloSheet(ExcelSheetStrategy):
    """
    Genera la hoja de análisis de trabajo paralelo por instancia para el reporte Excel.
    """

    def create_sheet(self, wb: Workbook, **kwargs: Any) -> None:
        all_results: List[Dict[str, Any]] = kwargs.get("all_results", [])
        
        ws = wb.create_sheet("Trabajo Paralelo")
        ws['A1'] = "ANÁLISIS DE TRABAJO PARALELO POR INSTANCIA"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        resultados_paralelos = []
        for res in all_results:
            inst_id = res.get('Instancia ID')
            if inst_id and inst_id != 'N/A' and inst_id != 'Principal':
                resultados_paralelos.append(res)
                
        if not resultados_paralelos:
            ws['A3'] = "No se detectó trabajo paralelo (instancias múltiples) en esta simulación."
            ws['A3'].font = Font(italic=True)
            return

        instancias = defaultdict(list)
        for res in resultados_paralelos:
            inst_id = res['Instancia ID']
            instancias[inst_id].append(res)
            
        headers = ["Tarea", "Instancia ID (8 dígitos)", "Trabajadores en Instancia", "Unidad Inicial", "Unidad Final", "Inicio", "Fin", "Duración Total (min)"]
        current_row = 3
        ws.append(headers)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        for inst_id, registros in sorted(instancias.items(), key=lambda item: (item[1][0].get('Tarea', ''), item[0])):
            registros.sort(key=lambda x: x.get('Inicio', datetime.min))
            tarea = registros[0].get('Tarea', 'N/A')
            trabajadores = registros[0].get('Lista Trabajadores', [])
            unidad_inicial = min((r.get('Numero Unidad', 0) for r in registros), default=0)
            unidad_final = max((r.get('Numero Unidad', 0) for r in registros), default=0)
            from typing import cast
            inicio_list = [cast(datetime, r.get('Inicio')) for r in registros if isinstance(r.get('Inicio'), datetime)]
            inicio = min(inicio_list) if inicio_list else None
            fin_list = [cast(datetime, r.get('Fin')) for r in registros if isinstance(r.get('Fin'), datetime)]
            fin = max(fin_list) if fin_list else None
            duracion_total = sum(r.get('Duracion (min)', 0) for r in registros)
            
            row_data = [
                tarea, 
                inst_id[:8], 
                ", ".join(trabajadores), 
                unidad_inicial, 
                unidad_final, 
                inicio.strftime('%d/%m/%Y %H:%M') if inicio else 'N/A', 
                fin.strftime('%d/%m/%Y %H:%M') if fin else 'N/A', 
                round(duracion_total, 2)
            ]
            ws.append(row_data)
            
            if (current_row - 3) % 2 == 0:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1

        column_widths = [30, 18, 30, 12, 12, 20, 20, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
        ws.freeze_panes = "A4"

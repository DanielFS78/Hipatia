"""
Nombre del Módulo: report_sheets.cronograma
Descripcion: Hoja Excel cronológica detallada por unidad/tarea de producción.
"""

from typing import List, Dict, Any
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .base import ExcelSheetStrategy

class CronogramaSheet(ExcelSheetStrategy):
    def create_sheet(self, wb: Workbook, all_results: List[Dict[str, Any]], hay_limite: bool = False, total_original: int | None = None, **kwargs: Any) -> None:  # type: ignore[override]
        ws = wb.create_sheet("Cronograma Detallado")
        ws['A1'] = "CRONOGRAMA DETALLADO POR UNIDAD Y ORDEN CRONOLÓGICO"
        ws['A1'].font, ws['A1'].fill = Font(size=14, bold=True, color="FFFFFF"), PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        ws.merge_cells('A1:M1')

        header_row = 4 if hay_limite else 3
        if hay_limite: ws['A2'] = f"⚠️ Mostrando {len(all_results)} de {total_original}"; ws['A2'].font = Font(size=10, italic=True, color="FF0000"); ws.merge_cells('A2:M2')

        headers = ["#", "Inicio", "Fin", "Tarea", "Instancia", "Grupo Trabajo", "Trabajador(es)", "Máquina", "Duración (min)", "Producto", "Unidad #", "Departamento", "Fab ID"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(header_row, col); c.value, c.font, c.fill, c.alignment = h, Font(bold=True, color="FFFFFF", size=10), PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"), Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row, last_date, seq = header_row + 1, None, 0
        grouped = defaultdict(list)
        for res in all_results: grouped[res.get('TareaDetalle', res.get('Tarea', 'N/A'))].append(res)

        for _, results in grouped.items():
            results.sort(key=lambda x: x.get('Inicio', datetime.min))
            p_ids = {r.get('Instancia ID', 'N/A') for r in results if r.get('Instancia ID', 'N/A') != 'N/A'}
            for task in results:
                start = task.get('Inicio')
                if not isinstance(start, datetime): continue
                if start.date() != last_date:
                    ws.merge_cells(f'A{current_row}:M{current_row}'); ws[f'A{current_row}'] = f"--- {start.strftime('%A, %d de %B de %Y')} ---"
                    ws[f'A{current_row}'].font, ws[f'A{current_row}'].fill, ws[f'A{current_row}'].alignment = Font(size=11, bold=True, color="FFFFFF"), PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"), Alignment(horizontal="center")
                    current_row += 1; last_date = start.date()

                seq += 1; inst = task.get('Instancia ID', 'N/A'); short = inst[:8] if inst != 'N/A' else 'Principal'
                trabs = task.get('Trabajador Asignado', 'N/A')
                fin = task.get('Fin')
                fin_str = fin.strftime('%H:%M') if isinstance(fin, datetime) else 'N/A'
                row_data = [seq, start.strftime('%d/%m %H:%M'), fin_str, task.get('Tarea', 'Sin nombre'), short, ", ".join(task.get('Lista Trabajadores', [])) if task.get('Lista Trabajadores') else 'N/A', ', '.join(trabs) if isinstance(trabs, list) else str(trabs), str(task.get('nombre_maquina', 'N/A')), task.get('Duracion (min)', 0), f"{task.get('Codigo Producto', '')} - {task.get('Descripcion Producto', '')}", task.get('Numero Unidad', '?'), str(task.get('Departamento', 'General')), str(task.get('fabricacion_id', 'N/A'))]
                
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(current_row, col); c.value = val; c.alignment = Alignment(horizontal="center" if col in [1, 2, 3, 9, 11] else "left", vertical="center", wrap_text=True)
                    c.border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
                
                if len(p_ids) > 1 and inst != 'N/A':
                    fill = PatternFill(start_color=['E8F4F8', 'F8E8F4', 'F4F8E8'][abs(hash(inst))%3], end_color=['E8F4F8', 'F8E8F4', 'F4F8E8'][abs(hash(inst))%3], fill_type='solid')
                    for ci in range(1, 14): ws.cell(current_row, ci).fill = fill
                current_row += 1
            for ci in range(1, 14): ws.cell(current_row-1, ci).border = Border(bottom=Side(style='medium', color='B0B0B0'))

        widths = {'A':5, 'B':18, 'C':10, 'D':35, 'E':12, 'F':25, 'G':25, 'H':15, 'I':10, 'J':30, 'K':10, 'L':15, 'M':12}
        for l, w in widths.items(): ws.column_dimensions[l].width = w
        ws.freeze_panes = f"A{header_row + 1}"; ws.auto_filter.ref = ws.dimensions

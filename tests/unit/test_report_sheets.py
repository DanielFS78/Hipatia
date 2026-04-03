# -*- coding: utf-8 -*-
"""Tests unitarios para hojas de reportes Excel (ResumenEjecutivo, AnalisisTrabajadores, etc.)."""
import pytest
from unittest.mock import MagicMock, ANY
from openpyxl import Workbook
from datetime import datetime, timedelta
from core.services.report_sheets import (
    ResumenEjecutivoSheet, AnalisisTrabajadoresSheet, CronogramaSheet, CuellosBotollaSheet, AuditSheet
)
from core.services.calculation_audit import CalculationDecision, DecisionStatus

pytestmark = pytest.mark.unit


class TestReportSheets:
    
    @pytest.fixture
    def wb(self):
        return Workbook()

    def test_resumen_ejecutivo_sheet(self, wb):
        strategy = ResumenEjecutivoSheet()
        analysis = {
            'start_time': datetime.now(),
            'end_time': datetime.now() + timedelta(hours=1),
            'total_duration_min': 100.0,
            'total_instancias_paralelas': 0,
            'max_instancias_simultaneas': 0
        }
        datos_informe = {
            'data': [{'Codigo Producto': 'P1', 'Numero Unidad': 10, 'Inicio': datetime.now(), 'Fin': datetime.now()}],
            'fab_info': 'Test Fab'
        }
        
        strategy.create_sheet(wb, analysis=analysis, datos_informe=datos_informe)
        
        assert "Resumen Ejecutivo" in wb.sheetnames
        ws = wb["Resumen Ejecutivo"]
        assert ws['A1'].value == "RESUMEN EJECUTIVO - ANÁLISIS DE PRODUCCIÓN"
        
        # Verify corrected unit calculation logic
        # Should find "Unidades Totales:" and check value
        found_units = False
        for row in ws.iter_rows(min_col=1, max_col=2):
            if row[0].value == "Unidades Totales:":
                assert str(row[1].value) == "10"
                found_units = True
                break
        assert found_units

    def test_analisis_trabajadores_sheet(self, wb):
        strategy = AnalisisTrabajadoresSheet()
        all_results = [
            {'Trabajador Asignado': ['W1'], 'Duracion (min)': 60, 'Tarea': 'T1'},
            {'Trabajador Asignado': 'W2', 'Duracion (min)': 30, 'Tarea': 'T2'}
        ]
        
        strategy.create_sheet(wb, all_results=all_results)
        
        assert "Análisis Trabajadores" in wb.sheetnames
        ws = wb["Análisis Trabajadores"]
        assert ws['A1'].value == "ANÁLISIS POR TRABAJADOR"
        
        # Check logic for multiple workers
        found_w1 = False
        found_w2 = False
        for row in ws.iter_rows(min_row=4, max_row=10, min_col=1, max_col=1):
            if row[0].value == "W1": found_w1 = True
            if row[0].value == "W2": found_w2 = True
        assert found_w1
        assert found_w2

    def test_cronograma_sheet(self, wb):
        strategy = CronogramaSheet()
        all_results = [
            {'Tarea': 'T1', 'Inicio': datetime.now(), 'Fin': datetime.now(), 'Duracion (min)': 10, 'Instancia ID': 'INST1'}
        ]
        
        strategy.create_sheet(wb, all_results=all_results)
        
        assert "Cronograma Detallado" in wb.sheetnames
        ws = wb["Cronograma Detallado"]
        
        # Verify Instancia column exists and is populated
        # Headers are in row 3
        assert ws['E3'].value == "Instancia"
        # Data in row 5 (header=3, date separator=4, data=5)
        # Actually logic inserts date separator first.
        # Let's just find the data row
        found_inst = False
        for row in ws.iter_rows(min_row=5):
            if row[3].value == "T1": # Column D is Tarea
                assert row[4].value == "INST1"[:8] # Column E is Instancia (shortened)
                found_inst = True
                break
        assert found_inst

    def test_cuellos_botella_sheet(self, wb):
        strategy = CuellosBotollaSheet()
        audit_log = [
            CalculationDecision(
                timestamp=datetime.now(),
                decision_type="TIEMPO_INACTIVO",
                reason="Test reason",
                user_friendly_reason="Friendly reason",
                task_name="T1",
                status=DecisionStatus.WARNING,
                details={'wait_time': 10, 'trabajador': 'W1'}
            )
        ]
        
        strategy.create_sheet(wb, audit_log=audit_log, all_results=[])
        
        assert "Cuellos de Botella" in wb.sheetnames
        ws = wb["Cuellos de Botella"]
        assert "ANÁLISIS ULTRA DETALLADO" in ws['A1'].value

    def test_audit_sheet(self, wb):
        strategy = AuditSheet()
        audit_log = [
            CalculationDecision(
                timestamp=datetime.now(),
                decision_type="INFO",
                reason="Info reason",
                user_friendly_reason="Friendly info",
                task_name="T1",
                status=DecisionStatus.NEUTRAL
            ),
            # This one should be included because status is CRITICAL
            CalculationDecision(
                timestamp=datetime.now(),
                decision_type="ERROR",
                reason="Error reason",
                user_friendly_reason="Friendly error",
                task_name="T2",
                status=DecisionStatus.CRITICAL
            )
        ]
        
        strategy.create_sheet(wb, audit_log=audit_log)
        
        assert "Audit Log" in wb.sheetnames
        ws = wb["Audit Log"]
        
        # Verify filtering: T2 should be there
        found_t2 = False
        for row in ws.iter_rows(min_row=4, max_col=2):
            if row[1].value == "T2": found_t2 = True
        assert found_t2
